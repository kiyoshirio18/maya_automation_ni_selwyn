from pathlib import Path
from io import BytesIO
from datetime import datetime, timedelta
import streamlit as st
import tempfile
import pandas as pd
import polars as pl
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Directory helpers - FIXED: Use raw strings or double backslashes correctly
ACCOUNT_JOURNEY_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ACCOUNT JOURNEY"
MERGE_ACCOUNTS_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\MERGED ACCOUNTS\2026"
REMARK_REPORT_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\REMARK REPORT"

def get_full_path(directory: str, filename: str) -> str:
    """Combine directory and filename into full path"""
    return str(Path(directory) / filename)

def get_yesterday_date_formatted():
    yesterday = datetime.now() - timedelta(days=2)
    return f"{yesterday.strftime('%B')} {yesterday.day}, {yesterday.year}".upper()

def get_today_date_formatted():
    today = datetime.now()
    return f"{today.strftime('%B')} {today.day}, {today.year}".upper()

def get_yesterday_mmddyy():
    yesterday = datetime.now() - timedelta(days=1)
    return yesterday.strftime("%m%d%y")

# File helpers
def get_latest_file(directory, pattern):
    base = Path(directory)
    if not base.exists():
        return None

    files = list(base.glob(pattern))
    if not files:
        return None

    return max(files, key=lambda f: f.stat().st_mtime)

def get_latest_maj_file():
    return get_latest_file(
        ACCOUNT_JOURNEY_DIR,
        "maya_account_journey_*.xlsx"
    )

def get_latest_mma_file():
    latest = get_latest_file(
        MERGE_ACCOUNTS_DIR,
        "maya_merged_accounts_*.xlsx"
    )
    return latest.name if latest else ""

def get_latest_remark_report_file():
    latest = get_latest_file(
        REMARK_REPORT_DIR,
        "*.xlsx"
    )
    return latest.name if latest else ""

# ============================================================================
# FILE UTILITIES
# ============================================================================

def resolve_server_file(server_dir: str, server_input: str) -> tuple[Path | None, str | None]:
    """Resolve a server file by full path, relative path, or filename search."""
    base_dir = Path(server_dir)
    if not base_dir.exists() or not base_dir.is_dir():
        return None, f"Server folder is not reachable: {base_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Please provide a server file name or path."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)
    
    if requested_path.exists() and requested_path.is_file():
        return requested_path, None
    
    candidate = base_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None
    
    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid file name or relative path."
    
    matches = [p for p in base_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, f"Multiple files named '{requested_name}' were found. Please use a relative path.\nMatches:\n{preview}"
    
    return None, f"Server file not found from input: {requested}"
        
def read_excel_ws(file_path, sheet_name=None):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    data = list(ws.values)
    headers = data[0]
    rows = data[1:]

    return wb, ws, headers, rows

def append_rows_to_existing_excel(file_path, new_rows_df):
    """Append rows to existing Excel without changing any formatting."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Append only the new rows (not the full dataframe)
    for row in dataframe_to_rows(new_rows_df, index=False, header=False):
        ws.append(row)

    wb.save(file_path)

def save_workbook(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def concat_df(excel_files, schema):
    """Concatenate multiple Excel files into a single polars DataFrame."""
    excel_list = []

    def normalize_header(value):
        return "" if value is None else "".join(ch for ch in str(value).upper().strip() if ch.isalnum())

    normalized_schema = {normalize_header(name): name for name in schema.keys()}

    for uploaded_file in excel_files:
        df = pl.read_excel(uploaded_file)

        # Rename variant headers to the canonical schema names before type casting.
        rename_map = {}
        for col in df.columns:
            normalized_col = normalize_header(col)
            if normalized_col in normalized_schema:
                canonical_name = normalized_schema[normalized_col]
                if col != canonical_name:
                    rename_map[col] = canonical_name

        if rename_map:
            df = df.rename(rename_map)

        # Apply types only to columns that actually exist in the file.
        available_overrides = {
            col_name: col_type
            for col_name, col_type in schema.items()
            if col_name in df.columns
        }
        if available_overrides:
            df = df.with_columns(
                [
                    pl.col(col_name).cast(col_type, strict=False)
                    for col_name, col_type in available_overrides.items()
                ]
            )

        excel_list.append(df)
    if not excel_list:
        return None
    return pl.concat(excel_list, how="vertical")

daily_remark_schema = {
    "S.No": pl.Int64,
    "Date": pl.Date,
    "Time": pl.Datetime,
    "Debtor": pl.Utf8,
    "Account No.": pl.Int64,
    "Card No.": pl.Utf8,
    "Service No.": pl.Utf8,
    "DPD": pl.Int64,
    "Reason For Default": pl.Utf8,
    "Call Status": pl.Utf8,
    "Status": pl.Utf8,
    "Remark": pl.Utf8,
    "Remark By": pl.Utf8,
    "Remark Type": pl.Utf8,
    "Field Visit Date": pl.Utf8,
    "Collector": pl.Utf8,
    "Client": pl.Utf8,
    "Product Description": pl.Utf8,
    "Product Type": pl.Utf8,
    "Batch No": pl.Utf8,
    "Account Type": pl.Utf8,
    "Relation": pl.Utf8,
    "PTP Amount": pl.Float64,
    "Next Call": pl.Utf8,
    "PTP Date": pl.Utf8,
    "Claim Paid Amount": pl.Float64,
    "Claim Paid Date": pl.Utf8,
    "Dialed Number": pl.Utf8,
    "Days Past Write Off": pl.Int64,
    "Balance": pl.Float64,
    "Contact Type": pl.Utf8,
    "Call Duration": pl.Int64,
    "Talk Time Duration": pl.Int64
}

# ============================================================================
# UI SIDE
# ============================================================================

def automate_account_journey_update(maj_path, mma_path):
    try:
        from openpyxl import load_workbook
        import pandas as pd
        from copy import copy

        FIXED_RECEIVED_DATE = (datetime.now() - timedelta(days=1)).strftime("%m/%d/%Y")

        def normalize_header(value):
            return "".join(ch for ch in str(value).upper().strip() if ch.isalnum())

        def pick_column_index(headers, candidates):
            normalized_headers = [normalize_header(header) for header in headers]
            for candidate in candidates:
                normalized_candidate = normalize_header(candidate)
                if normalized_candidate in normalized_headers:
                    return normalized_headers.index(normalized_candidate)
            return None

        def looks_like_phone(text):
            raw = str(text).strip()
            digits = "".join(ch for ch in raw if ch.isdigit())
            return len(digits) >= 7 and len(digits) <= 15 and all(ch.isdigit() or ch in "+()- ." for ch in raw) and "." not in raw and "," not in raw

        def parse_balance_series(series):
            def _convert(value):
                if value is None:
                    return None
                text = str(value).strip()
                if text == "" or text.lower() == "nan" or looks_like_phone(text):
                    return None
                cleaned = text.replace(",", "").replace("$", "").replace("₱", "").strip()
                if cleaned.startswith("(") and cleaned.endswith(")"):
                    cleaned = "-" + cleaned[1:-1]
                try:
                    return float(cleaned)
                except Exception:
                    return None

            return series.apply(_convert)

        def parse_date_series(series):
            return pd.to_datetime(series, errors="coerce", dayfirst=False)

        def format_date_series(series):
            return series.dt.strftime("%m/%d/%Y")

        # Load MAJ workbook (preserves all formatting)
        wb = load_workbook(maj_path)
        ws = wb.active

        # Read MAJ headers (first row)
        maj_headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[1]
        ]

        maj_placement_idx = pick_column_index(maj_headers, ["PLACEMENT"])
        maj_account_idx = pick_column_index(maj_headers, ["ACCOUNT", "ACCOUNT NUMBER", "ACCOUNT_NO", "ACCOUNT NUMBER/NAME"])
        maj_endo_date_idx = pick_column_index(maj_headers, ["ENDO DATE", "ENDO_DATE", "RECEIVED DATE", "RECEIVED_DATE"])
        maj_balance_idx = pick_column_index(maj_headers, ["BALANCE", "OB", "OUTSTANDING BALANCE", "OUTSTANDING_BALANCE", "OSB"])
        maj_product_idx = pick_column_index(maj_headers, ["PRODUCT NAME", "PRODUCT_NAME"])

        if maj_account_idx is None:
            maj_account_idx = 1

        if maj_endo_date_idx is None or maj_balance_idx is None:
            return None, "MAJ does not contain the expected 'Endo Date' and 'Balance' columns."

        # Build a set of existing values from the actual MAJ account column
        existing_accounts = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > maj_account_idx:
                val = row[maj_account_idx]
                if val not in (None, ""):
                    existing_accounts.add(str(val).strip())

        # Load MMA into DataFrame (preserve order of columns)
        mma_df = pd.read_excel(mma_path, header=0, dtype=str)
        mma_df = mma_df.fillna("")

        # Locate MMA important columns using exact normalized headers.
        mma_placement_idx = pick_column_index(mma_df.columns, ["PLACEMENT"])
        mma_account_idx = pick_column_index(mma_df.columns, ["ACCOUNT", "ACCOUNT NUMBER", "ACCOUNT_NO", "ACCOUNT NUMBER/NAME"])
        mma_received_date_idx = pick_column_index(mma_df.columns, ["RECEIVED DATE", "RECEIVED_DATE"])
        mma_ob_idx = pick_column_index(mma_df.columns, ["OB", "BALANCE", "OUTSTANDING BALANCE", "OUTSTANDING_BALANCE", "OSB"])

        if mma_placement_idx is None or mma_account_idx is None:
            return None, "MMA does not contain the expected 'Placement' and 'Account' columns."
        if mma_received_date_idx is None or mma_ob_idx is None:
            return None, "MMA does not contain the expected 'RECEIVED DATE' and 'OB' columns."

        placement_series = mma_df.iloc[:, mma_placement_idx].astype(str).str.strip()
        account_series = mma_df.iloc[:, mma_account_idx].astype(str).str.strip()
        received_series = mma_df.iloc[:, mma_received_date_idx]
        ob_series = mma_df.iloc[:, mma_ob_idx]

        received_parsed = parse_date_series(received_series)
        received_text = format_date_series(received_parsed)
        ob_values = parse_balance_series(ob_series)

        fixed_date_parsed = pd.to_datetime(FIXED_RECEIVED_DATE, format="%m/%d/%Y", errors="coerce")
        if pd.isna(fixed_date_parsed):
            return None, f"Invalid fixed received date configured: {FIXED_RECEIVED_DATE}"

        fixed_date_text = fixed_date_parsed.strftime("%m/%d/%Y")
        date_filter = received_text.fillna("") == fixed_date_text

        # Create a 'name match' series: check whether MMA account exists in MAJ column B
        name_match = account_series.apply(lambda x: x if x in existing_accounts else "#N/A")

        # Create a preview dataframe with 'name match' as first column (in-memory)
        preview_df = pd.DataFrame(
            {
                "name match": name_match,
                "placement": placement_series,
                "account": account_series,
                "received_date": received_text,
                "ob": ob_values,
                "product_name": placement_series,
            }
        )

        # Count MAJ rows before appending (based on non-empty account numbers).
        maj_rows_before = 0
        for row_idx in range(2, ws.max_row + 1):
            acct_val = ws.cell(row=row_idx, column=maj_account_idx + 1).value
            if acct_val not in (None, ""):
                maj_rows_before += 1

        # Filter rows that match the fixed received date and do not already exist in MAJ.
        filtered_idx = (preview_df["name match"] == "#N/A") & date_filter
        filtered_count = filtered_idx.sum()
        total_rows = len(preview_df)
        matched_count = total_rows - filtered_count

        if filtered_count == 0:
            kpi_summary = {
                "maj_rows_before": int(maj_rows_before),
                "added_from_mma": 0,
                "maj_rows_after": int(maj_rows_before),
                "added_start_row": None,
                "added_end_row": None,
            }
            return wb, f"No rows appended: no MMA rows matched received date {fixed_date_text} with unmatched accounts (matched={matched_count}, unmatched=0).", kpi_summary

        # Prepare to append rows to MAJ.
        needed_cols = max(
            len(maj_headers),
            maj_endo_date_idx + 1,
            maj_balance_idx + 1,
            maj_account_idx + 1,
            (maj_product_idx + 1) if maj_product_idx is not None else 0,
        )

        # Find the last real data row using key MAJ columns (A-F) instead of ws.max_row,
        # because formatted templates can make append land far below visible data.
        last_data_row = 1
        for row_idx in range(2, ws.max_row + 1):
            has_data = any(
                ws.cell(row=row_idx, column=col_idx).value not in (None, "")
                for col_idx in range(1, min(needed_cols, 6) + 1)
            )
            if has_data:
                last_data_row = row_idx

        insert_row = last_data_row + 1

        # Append filtered rows mapping MMA -> MAJ
        for i in preview_df[filtered_idx].index:
            new_row = [None] * needed_cols

            if maj_placement_idx is not None:
                new_row[maj_placement_idx] = placement_series.iloc[i]
            new_row[maj_account_idx] = account_series.iloc[i]
            received_value = received_parsed.iloc[i]
            if pd.isna(received_value):
                received_value = fixed_date_parsed
            new_row[maj_endo_date_idx] = received_value.to_pydatetime()
            new_row[maj_balance_idx] = ob_values.iloc[i]
            if maj_product_idx is not None:
                new_row[maj_product_idx] = placement_series.iloc[i] if str(placement_series.iloc[i]).strip() else None

            # Copy the formatting from the row above, similar to using Format Painter.
            template_row = insert_row - 1 if insert_row > 1 else last_data_row
            if template_row >= 1:
                for col_idx in range(1, needed_cols + 1):
                    source_cell = ws.cell(row=template_row, column=col_idx)
                    target_cell = ws.cell(row=insert_row, column=col_idx)
                    target_cell._style = copy(source_cell._style)
                    target_cell.number_format = source_cell.number_format
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.protection = copy(source_cell.protection)

            for col_idx, value in enumerate(new_row, start=1):
                ws.cell(row=insert_row, column=col_idx, value=value)

            insert_row += 1

        start_row = last_data_row + 1
        end_row = insert_row - 1
        kpi_summary = {
            "maj_rows_before": int(maj_rows_before),
            "added_from_mma": int(filtered_count),
            "maj_rows_after": int(maj_rows_before + filtered_count),
            "added_start_row": int(start_row),
            "added_end_row": int(end_row),
        }
        return wb, f"✅ Appended {filtered_count} MMA row(s) for received date {fixed_date_text} to MAJ (matched={matched_count}, total={total_rows}, rows={start_row}-{end_row}).", kpi_summary

    except Exception as e:
        import traceback
        return None, f"Error: {str(e)}\n{traceback.format_exc()}", None

# STREAMLIT UI - Only Account Journey Update
st.set_page_config(page_title="Account Journey Automation", layout="wide")
st.title("Automation for ACCOUNT JOURNEY")
st.caption("Complete Account Journey Update")

if "processed_data" not in st.session_state:
    st.session_state.processed_data = None

if "maj_download_bytes" not in st.session_state:
    st.session_state.maj_download_bytes = None
if "maj_download_filename" not in st.session_state:
    st.session_state.maj_download_filename = None
if "dca_download_bytes" not in st.session_state:
    st.session_state.dca_download_bytes = None
if "dca_download_filename" not in st.session_state:
    st.session_state.dca_download_filename = None
if "acc_total_rows_before" not in st.session_state:
    st.session_state.acc_total_rows_before = None
if "acc_total_added_from_mma" not in st.session_state:
    st.session_state.acc_total_added_from_mma = None
if "acc_total_pulled_out" not in st.session_state:
    st.session_state.acc_total_pulled_out = 0
if "acc_total_rows_after" not in st.session_state:
    st.session_state.acc_total_rows_after = None


tab1 = st.container()

with tab1:
    st.markdown("### Main Account Journey Update")

    with st.form("main_update_form"):
        default_maj_filename = f"maya_account_journey_{get_yesterday_date_formatted()}.xlsx"
        default_mma_filename = get_latest_mma_file() or f"maya_merged_accounts_{get_yesterday_mmddyy()}.xlsx"

        maj_path = st.text_input(
            "MAJ File (maya_account_journey_*.xlsx)",
            value=default_maj_filename,
            help=f"Full path: {ACCOUNT_JOURNEY_DIR}\n\nAuto-populated with yesterday's date: {get_yesterday_date_formatted()}"
        )
        uploaded_maj = st.file_uploader("Or upload MAJ file", type=["xlsx"], key="main_upload_maj")

        mma_path = st.text_input(
            "MMA File (maya_merged_accounts_*.xlsx)",
            value=default_mma_filename,
            help=f"Full path: {MERGE_ACCOUNTS_DIR}\n\nAuto-suggests the latest file in the Merge Accounts directory or yesterday's mma filename."
        )
        uploaded_mma = st.file_uploader("Or upload MMA file", type=["xlsx"], key="main_upload_mma")

        default_remark_filename = get_latest_remark_report_file() or ""
        remark_path = st.text_input(
            "Daily Remark Report (for Account Journey)",
            value=default_remark_filename,
            help=f"Full path: {REMARK_REPORT_DIR}\n\nAuto-suggests the latest file in the Remark Report directory."
        )
        uploaded_remark = st.file_uploader("Or upload Daily Remark Report file", type=["xlsx"], accept_multiple_files=True, key="uploaded_remark")

        submitted_main = st.form_submit_button("Submit", use_container_width=True, type="primary")

    if submitted_main:
        st.session_state.maj_download_bytes = None
        st.session_state.maj_download_filename = None
        st.session_state.dca_download_bytes = None
        st.session_state.dca_download_filename = None

        maj_source = uploaded_maj if uploaded_maj else maj_path
        mma_source = uploaded_mma if uploaded_mma else mma_path
        
        if not maj_source or not mma_source:
            st.error("Please provide both Maya Account Journey and Maya Merged Accounts files.")
        else:
            try:
                temp_files = []
                
                # ----------------------------
                # HANDLE MAJ INPUT
                # ----------------------------
                if uploaded_maj:
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    tmp.write(uploaded_maj.getvalue())
                    tmp.close()
                    maj_source = tmp.name
                    temp_files.append(tmp.name)
                else:
                    if '\\' not in str(maj_source) and '/' not in str(maj_source):
                        candidate = Path(ACCOUNT_JOURNEY_DIR) / maj_source
                        if candidate.exists():
                            maj_source = str(candidate)
                        else:
                            maj_name = Path(maj_source).name
                            matches = list(Path(ACCOUNT_JOURNEY_DIR).rglob(maj_name))
                            if matches:
                                maj_source = str(max(matches, key=lambda p: p.stat().st_mtime))
                # ----------------------------
                # HANDLE MMA INPUT
                # ----------------------------
                if uploaded_mma:
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    tmp.write(uploaded_mma.getvalue())
                    tmp.close()
                    mma_source = tmp.name
                    temp_files.append(tmp.name)
                else:
                    if '\\' not in str(mma_source) and '/' not in str(mma_source):
                        candidate = Path(MERGE_ACCOUNTS_DIR) / mma_source
                        if candidate.exists():
                            mma_source = str(candidate)
                        else:
                            mma_name = Path(mma_source).name
                            matches = list(Path(MERGE_ACCOUNTS_DIR).rglob(mma_name))
                            if matches:
                                mma_source = str(max(matches, key=lambda p: p.stat().st_mtime))
                # ----------------------------
                # VALIDATE FILE EXISTENCE
                # ----------------------------
                if not uploaded_maj and not Path(maj_source).exists():
                    st.error(f"❌ MAJ file not found: {maj_source}")
                    st.stop()
                
                if not uploaded_mma and not Path(mma_source).exists():
                    st.error(f"❌ MMA file not found: {mma_source}")
                    st.stop()
                
                # ----------------------------
                # RUN FILTER LOGIC
                # ----------------------------
                with st.spinner("Running Match & Filter..."):
                    wb, result_message, kpi_summary = automate_account_journey_update(
                        maj_source, mma_source
                    )

                    if kpi_summary:
                        st.session_state.acc_total_rows_before = int(kpi_summary.get("maj_rows_before", 0))
                        st.session_state.acc_total_added_from_mma = int(kpi_summary.get("added_from_mma", 0))
                        st.session_state.acc_total_rows_after = int(kpi_summary.get("maj_rows_after", 0))
                        st.session_state.acc_total_pulled_out = 0

                    if wb is not None:
                        # Download (format-safe)
                        output = save_workbook(wb)

                        filename = f"maya_account_journey_{get_today_date_formatted()}.xlsx"
                        st.session_state.maj_download_bytes = output.getvalue()
                        st.session_state.maj_download_filename = filename
                    else:
                        st.error(result_message)
                    # If Daily Remark files were provided in the main form, also generate Account Journey
                    try:
                        acc_files = uploaded_remark if uploaded_remark else []
                        remark_source = uploaded_remark if uploaded_remark else None
                        
                        # If no upload but a directory path was provided, resolve it
                        if not remark_source and remark_path:
                            if '\\' not in str(remark_path) and '/' not in str(remark_path):
                                candidate = Path(REMARK_REPORT_DIR) / remark_path
                                if candidate.exists():
                                    remark_source = [candidate]
                                else:
                                    remark_name = Path(remark_path).name
                                    matches = list(Path(REMARK_REPORT_DIR).rglob(remark_name))
                                    if matches:
                                        remark_source = [max(matches, key=lambda p: p.stat().st_mtime)]
                            else:
                                remark_path_obj = Path(remark_path)
                                if remark_path_obj.exists():
                                    remark_source = [remark_path_obj]
                        
                        acc_files = remark_source if remark_source else []
                        if acc_files:
                            with st.spinner("Generating Account Journey..."):
                                try:
                                    maya_dispositions = pl.read_csv("./resources/maya_dispositions.csv")
                                    
                                    # Concatenate all Daily Remark files using concat_df
                                    daily_remark = concat_df(acc_files, daily_remark_schema)
                                    if daily_remark is not None:
                                        account_journey = daily_remark.join(maya_dispositions, left_on="Status", right_on="VOLARE STATUS", how="left")
                                        account_journey = account_journey.with_columns(pl.col("HIERARCHY").fill_null(0)).sort("HIERARCHY", descending=True)
                                        account_journey = account_journey.unique(subset=["Date", "Account No."], keep="first", maintain_order=True)

                                        out_bytes = BytesIO()
                                        account_journey.write_excel(out_bytes, worksheet="ACC JOURNEY", dtype_formats={pl.Int64: "0", pl.Date: "mm/dd/yyyy"})
                                        out_bytes.seek(0)

                                        today_name = datetime.now().strftime("%B %d, %Y").upper()

                                        dca_filename = f"account_journey_dca_{today_name}.xlsx"
                                        maj_filename = f"maya_account_journey_matched_{today_name}.xlsx"

                                        # Perform matching: map Account No. + Date in Account Journey -> PROPOSED DISPOSITION
                                        try:
                                            def normalize_header(v):
                                                return "" if v is None else "".join(ch for ch in str(v).upper().strip() if ch.isalnum())

                                            def normalize_account(value):
                                                if value is None:
                                                    return ""
                                                if isinstance(value, int):
                                                    return str(value)
                                                if isinstance(value, float):
                                                    if pd.isna(value):
                                                        return ""
                                                    if float(value).is_integer():
                                                        return str(int(value))
                                                    return str(value).strip()
                                                text = str(value).strip()
                                                if text.endswith(".0") and text[:-2].isdigit():
                                                    return text[:-2]
                                                return text

                                            def date_key_candidates(value):
                                                if value is None:
                                                    return set()
                                                text = str(value).strip()
                                                if text == "":
                                                    return set()

                                                keys = set()
                                                for dayfirst in (False, True):
                                                    parsed = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
                                                    if not pd.isna(parsed):
                                                        keys.add(parsed.date().isoformat())
                                                        keys.add(parsed.strftime("%m/%d/%Y"))
                                                return keys

                                            # locate candidate column names in polars dataframe
                                            def find_polars_col(df, candidates):
                                                norm = {normalize_header(c): c for c in df.columns}
                                                for cand in candidates:
                                                    n = normalize_header(cand)
                                                    if n in norm:
                                                        return norm[n]
                                                return None

                                            acct_col_name = find_polars_col(account_journey, ["Account No.", "Account No", "Account", "ACCOUNT NO", "ACCOUNT"])
                                            date_col_name = find_polars_col(account_journey, ["Date"])
                                            disposition_col_name = find_polars_col(account_journey, ["PROPOSED DISPOSITION", "PROPOSED_DISPOSITION", "PROPOSED DISPOSITION"])

                                            mapping = {}
                                            mapping_by_account = {}
                                            if acct_col_name and date_col_name and disposition_col_name:
                                                acct_vals = account_journey[acct_col_name].to_list()
                                                date_vals = account_journey[date_col_name].to_list()
                                                disposition_vals = account_journey[disposition_col_name].to_list()
                                                for a, d, disp in zip(acct_vals, date_vals, disposition_vals):
                                                    if a is None:
                                                        continue
                                                    account_key = normalize_account(a)
                                                    if account_key == "":
                                                        continue
                                                    for date_key in date_key_candidates(d):
                                                        # normalize disposition: treat NaN/blank as missing so fallback applies
                                                        if pd.isna(disp):
                                                            disp_norm = ""
                                                        else:
                                                            disp_norm = str(disp).strip()
                                                            if disp_norm.lower() == "nan":
                                                                disp_norm = ""

                                                        mapping[(account_key, date_key)] = disp_norm
                                                        mapping_by_account.setdefault(account_key, {})[date_key] = disp_norm
                                        except Exception:
                                            mapping = {}
                                            mapping_by_account = {}

                                        # Add ACC JOURNEY sheet to MAJ workbook and write matched values into the correct date columns
                                        try:
                                            if wb is not None:
                                                # Convert polars to pandas and add as sheet to wb
                                                acc_pd = account_journey.to_pandas()
                                                if 'ACC JOURNEY' in wb.sheetnames:
                                                    wb.remove(wb['ACC JOURNEY'])
                                                ws_acc = wb.create_sheet('ACC JOURNEY')

                                                # Write header and rows
                                                for r_idx, r in enumerate(dataframe_to_rows(acc_pd, index=False, header=True), start=1):
                                                    for col_idx, val in enumerate(r, start=1):
                                                        ws_acc.cell(row=r_idx, column=col_idx, value=val)

                                                # Select the MAJ sheet (first sheet that's not ACC JOURNEY)
                                                ws = None
                                                for sheet_name in wb.sheetnames:
                                                    if sheet_name != 'ACC JOURNEY':
                                                        ws = wb[sheet_name]
                                                        break

                                                if ws is None:
                                                    ws = wb.active  # fallback

                                                maj_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

                                                def pick_column_index(headers, candidates):
                                                    normalized = [normalize_header(h) for h in headers]
                                                    for cand in candidates:
                                                        n = normalize_header(cand)
                                                        if n in normalized:
                                                            return normalized.index(n)
                                                    return None

                                                acct_idx = pick_column_index(maj_headers, ["ACCOUNT", "ACCOUNT NO", "ACCOUNT NO.", "ACCOUNT NUMBER"])

                                                # Build a map of MAJ date columns from the header row.
                                                date_column_indexes = {}
                                                date_column_order = []
                                                for col_idx, header_value in enumerate(maj_headers, start=1):
                                                    for date_key in date_key_candidates(header_value):
                                                        date_column_indexes[date_key] = col_idx
                                                        parsed_header_date = pd.to_datetime(date_key, errors="coerce")
                                                        if not pd.isna(parsed_header_date):
                                                            date_column_order.append((parsed_header_date.date(), col_idx))

                                                # Keep unique columns in date order so earlier columns can be marked correctly.
                                                seen_date_cols = set()
                                                date_column_order = [
                                                    (date_value, col_idx)
                                                    for date_value, col_idx in sorted(date_column_order, key=lambda item: (item[0], item[1]))
                                                    if not (col_idx in seen_date_cols or seen_date_cols.add(col_idx))
                                                ]

                                                # Find POUT DATE column in MAJ
                                                pout_date_col_idx = None
                                                for col_idx, header_value in enumerate(maj_headers, start=1):
                                                    if normalize_header(header_value) in [normalize_header(c) for c in ["POUT DATE", "POUT_DATE", "PULLED OUT DATE", "PULLED_OUT_DATE"]]:
                                                        pout_date_col_idx = col_idx
                                                        break

                                                # Load MMA data to extract POUT DATE mapping (account -> pout_date)
                                                pout_date_mapping = {}
                                                try:
                                                    # Resolve MMA source path
                                                    mma_source_resolved = mma_source
                                                    if isinstance(mma_source, str) and not Path(mma_source).exists():
                                                        mma_resolved_path, _ = resolve_server_file(MERGE_ACCOUNTS_DIR, mma_source)
                                                        if mma_resolved_path:
                                                            mma_source_resolved = str(mma_resolved_path)
                                                    
                                                    if Path(mma_source_resolved).exists():
                                                        mma_check_df = pd.read_excel(mma_source_resolved, header=0, dtype=str)
                                                        mma_acct_col = pick_column_index(mma_check_df.columns, ["ACCOUNT", "ACCOUNT NUMBER", "ACCOUNT_NO", "ACCOUNT NUMBER/NAME"])
                                                        mma_pout_col = pick_column_index(mma_check_df.columns, ["PULLED OUT DATE", "PULLED_OUT_DATE", "PULLOUT DATE", "PULLOUT_DATE", "POUT DATE"])
                                                        
                                                        if mma_acct_col is not None and mma_pout_col is not None:
                                                            mma_accounts = mma_check_df.iloc[:, mma_acct_col].astype(str).str.strip()
                                                            mma_pout_dates = mma_check_df.iloc[:, mma_pout_col]
                                                            mma_pout_parsed = pd.to_datetime(mma_pout_dates, errors="coerce")
                                                            
                                                            for account, pout_dt in zip(mma_accounts, mma_pout_parsed):
                                                                if account and not pd.isna(pout_dt):
                                                                    pout_date_mapping[account.strip()] = pout_dt.to_pydatetime()
                                                except Exception:
                                                    # POUT DATE mapping is optional
                                                    pass

                                                if acct_idx is not None and date_column_indexes:
                                                    # Calculate target date (yesterday, same as FIXED_RECEIVED_DATE)
                                                    fixed_date_dt = datetime.now() - timedelta(days=1)
                                                    fixed_date_str = fixed_date_dt.strftime("%m/%d/%Y")
                                                    fixed_date_iso = fixed_date_dt.date().isoformat()
                                                    fixed_date_keys = [fixed_date_str, fixed_date_iso]
                                                    
                                                    # Determine target date column (supports multiple date key formats)
                                                    target_date_col_idx = None
                                                    for date_key in fixed_date_keys:
                                                        if date_key in date_column_indexes:
                                                            target_date_col_idx = date_column_indexes[date_key]
                                                            break
                                                    
                                                    if target_date_col_idx is not None:
                                                        # Find last row with a non-empty account number
                                                        last_account_row = 1
                                                        for row_idx in range(2, ws.max_row + 1):
                                                            acc_val = ws.cell(row=row_idx, column=acct_idx + 1).value
                                                            if acc_val not in (None, ""):
                                                                last_account_row = row_idx

                                                        pulled_out_written_count = 0
                                                        
                                                        # Write disposition/NO ANSWER only to the target date column, up to last account row
                                                        for row_idx in range(2, last_account_row + 1):
                                                            acc_val = ws.cell(row=row_idx, column=acct_idx + 1).value
                                                            account_key = normalize_account(acc_val)
                                                            if account_key == "":
                                                                continue

                                                            # Check if account has a match for the target date
                                                            account_date_map = mapping_by_account.get(account_key, {})
                                                            mapped_disposition = ""
                                                            for date_key in fixed_date_keys:
                                                                if date_key in account_date_map:
                                                                    mapped_disposition = account_date_map.get(date_key, "")
                                                                    break
                                                            
                                                            # Write disposition if matched, otherwise "NO ANSWER FROM THE USER"
                                                            value_to_write = mapped_disposition if mapped_disposition not in (None, "") else "NO ANSWER FROM THE USER"
                                                            ws.cell(row=row_idx, column=target_date_col_idx, value=value_to_write)

                                                        # For rows added from MMA, mark all earlier date columns as not yet endorsed.
                                                        # Also copy font formatting from the previous column to the target date column for newly added rows.
                                                        added_start_row = None
                                                        added_end_row = None
                                                        if isinstance(kpi_summary, dict):
                                                            added_start_row = kpi_summary.get("added_start_row")
                                                            added_end_row = kpi_summary.get("added_end_row")

                                                        if added_start_row is not None and added_end_row is not None:
                                                            # Get font from the previous column (target_date_col_idx - 1) to copy to newly added rows
                                                            if target_date_col_idx > 1:
                                                                from copy import copy
                                                                source_font_col = target_date_col_idx - 1
                                                                template_font = ws.cell(row=int(added_start_row), column=source_font_col).font
                                                            else:
                                                                template_font = None

                                                            for row_idx in range(int(added_start_row), int(added_end_row) + 1):
                                                                for date_value, col_idx in date_column_order:
                                                                    if col_idx == target_date_col_idx:
                                                                        continue
                                                                    if date_value < fixed_date_dt.date():
                                                                        ws.cell(row=row_idx, column=col_idx, value="NOT YET ENDORSED")

                                                            # Apply font to the target date column for newly added rows
                                                            if template_font is not None:
                                                                from copy import copy
                                                                for row_idx in range(int(added_start_row), int(added_end_row) + 1):
                                                                    target_cell = ws.cell(row=row_idx, column=target_date_col_idx)
                                                                    target_cell.font = copy(template_font)

                                                        # Write POUT DATE match if column exists
                                                        if pout_date_col_idx is not None:
                                                            if added_start_row is not None and added_end_row is not None:
                                                                for row_idx in range(int(added_start_row), int(added_end_row) + 1):
                                                                    acc_val = ws.cell(row=row_idx, column=acct_idx + 1).value
                                                                    account_key = normalize_account(acc_val)
                                                                    if account_key == "" or account_key not in pout_date_mapping:
                                                                        continue
                                                                    ws.cell(row=row_idx, column=pout_date_col_idx, value=pout_date_mapping[account_key])
                                                                    pulled_out_written_count += 1

                                                        st.session_state.acc_total_pulled_out = int(pulled_out_written_count)

                                                        # Keep only the updated MAJ sheet in the exported workbook.
                                                        if "ACC JOURNEY" in wb.sheetnames:
                                                            wb.remove(wb["ACC JOURNEY"])

                                                    # offer the updated MAJ for download
                                                    updated_output = save_workbook(wb)
                                                    st.session_state.maj_download_bytes = updated_output.getvalue()
                                                    st.session_state.maj_download_filename = maj_filename
                                                    st.session_state.dca_download_bytes = out_bytes.getvalue()
                                                    st.session_state.dca_download_filename = dca_filename
                                        except Exception as exc:
                                            st.error(f"Failed to write matches to MAJ: {exc}")
                                except Exception as exc:
                                    st.error(f"Failed to generate Account Journey: {exc}")
                    except NameError:
                        # acc_files may not be defined if user never saw the uploader
                        pass
                
                # ----------------------------
                # CLEANUP TEMP FILES
                # ----------------------------
                for temp_file in temp_files:
                    try:
                        Path(temp_file).unlink(missing_ok=True)
                    except:
                        pass
                        
            except Exception as exc:
                st.error(f"❌ Automation failed: {exc}")
                st.write("Debug info - maj_source:", maj_source)
                st.write("Debug info - mma_source:", mma_source)

    if st.session_state.maj_download_bytes and st.session_state.maj_download_filename:
        st.download_button(
            "📥 Download Updated MAJ (with Account Journey matches)",
            data=st.session_state.maj_download_bytes,
            file_name=st.session_state.maj_download_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # Show KPI summary if available
    if st.session_state.get("acc_total_rows_before") is not None:
        st.subheader("Filter KPI")

        total_rows_before = st.session_state.get("acc_total_rows_before", 0)
        total_added_from_mma = st.session_state.get("acc_total_added_from_mma", 0)
        total_pulled_out = st.session_state.get("acc_total_pulled_out", 0)
        total_rows_after = st.session_state.get("acc_total_rows_after", 0)

        metric_columns = st.columns(4)

        metric_columns[0].metric(
            label="Total Rows (MAJ Before Mapping)",
            value=f"{total_rows_before:,}",
        )
        metric_columns[1].metric(
            label="Total Added Data (MMA -> MAJ)",
            value=f"{total_added_from_mma:,}",
        )
        metric_columns[2].metric(
            label="Total Pulled Out (POUT DATE)",
            value=f"{total_pulled_out:,}",
        )
        metric_columns[3].metric(
            label="Total Rows After Adding MMA",
            value=f"{total_rows_after:,}",
        )

    if st.session_state.dca_download_bytes and st.session_state.dca_download_filename:
        st.download_button(
            "📥 Download Account Journey / DCA",
            data=st.session_state.dca_download_bytes,
            file_name=st.session_state.dca_download_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown("---")
st.caption("Account Journey Automation Tool | v2.3")