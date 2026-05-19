import streamlit as st
import pandas as pd
import polars as pl
import msoffcrypto
import re
from pathlib import Path
from collections import defaultdict
from io import BytesIO
from datetime import date, datetime
from typing import Dict, Tuple
from openpyxl import load_workbook

st.header("Agency Metrics")

# ── Constants ────────────────────────────────────────────────────────────────
STATUS_COLUMN = "status"
CYCLE_COLUMN  = "cycle"

METRIC_LABELS = {
    "dials":            "Dials",
    "manual_dials":     "Manual Dials",
    "rpc_under_nego":   "RPC Under Nego",
    "rpc_niop":         "RPC NIOP",
    "third_party":      "3rd Party Contacted",
    "dispute":          "Dispute",
    "email_sent":       "Email Sent",
    "email_responsive": "Email Responsive",
    "vb_count":         "VB Count",
    "vb_connected":     "VB Connected",
}

USED_COLUMNS = {STATUS_COLUMN, CYCLE_COLUMN, "account_no.", "call_duration", "remark_type", "remark"}

DRR_CATEGORY_STATUS = {
    "UNDERNEGO": ["POSITIVE CONTACT - CALLBACK", "POSITIVE CONTACT - UNDERNEGO"],
    "NIOP": ["POSITIVE CONTACT - RPC REFUSE TO PAY", "POSITIVE CONTACT - DISPUTE"],
    "3RD PARTY CONTACTED": ["POSITIVE - 3RD PARTY CONTACTED"],
    "DISPUTE": ["POSITIVE CONTACT - DISPUTE"],
}

DEFAULT_LOCAL_OPTIONS_DIR = "options"
DEFAULT_ENDORSEMENT_FILE = f"MAYA ENDORSEMENT {date.today().strftime('%m%d%Y')}.xlsx"
DEFAULT_WORKBOOK_PASSWORD = "Maya@2026"
DEFAULT_SERVER_MASTERFILE_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ENDO\MASTERFILE"
DERIVED_DATE_STATE_KEY = "_agency_metrics_derived_date"


class LocalInputFile:
    def __init__(self, file_path: Path):
        self.path = file_path
        self.name = file_path.name

    def getvalue(self) -> bytes:
        return self.path.read_bytes()


def resolve_server_endorsement_file(server_input: str) -> tuple[Path | None, str | None]:
    master_dir = Path(DEFAULT_SERVER_MASTERFILE_DIR)
    if not master_dir.exists() or not master_dir.is_dir():
        return None, f"Server folder is not reachable: {master_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Server file mode is enabled but no file name/path was provided."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)

    if requested_path.exists() and requested_path.is_file():
        return requested_path, None

    candidate = master_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None

    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid server file name or relative path."

    matches = [
        p for p in master_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()
    ]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, (
            f"Multiple files named '{requested_name}' were found under MASTERFILE. "
            "Please paste a relative path, for example 'MARCH 2026\\filename.xlsx'.\n"
            f"Matches:\n{preview}"
        )

    return None, f"Server file not found from input: {requested}"


def get_local_endorsement_files() -> list[LocalInputFile]:
    script_dir = Path(__file__).resolve().parent
    project_dir = script_dir.parent

    options_dir = project_dir / DEFAULT_LOCAL_OPTIONS_DIR
    scan_dirs = []
    if options_dir.exists() and options_dir.is_dir():
        scan_dirs.append(options_dir)
    scan_dirs.append(project_dir)

    all_files: list[Path] = []
    for base_dir in scan_dirs:
        all_files.extend(
            [
                p
                for p in base_dir.rglob("*")
                if p.is_file()
                and p.suffix.lower() in {".xlsx", ".xls", ".xlsb", ".csv"}
                and not p.name.startswith("~$")
                and not p.name.startswith("~")
            ]
        )

    endorsement = [
        LocalInputFile(p)
        for p in all_files
        if "endorsement" in p.name.lower()
    ]
    endorsement.sort(
        key=lambda file_obj: (
            file_obj.name.lower() != DEFAULT_ENDORSEMENT_FILE.lower(),
            file_obj.name.lower(),
        )
    )
    return endorsement


def make_unique_columns(columns: list) -> list[str]:
    seen = {}
    unique = []
    for raw_name in columns:
        base = str(raw_name).strip() if raw_name is not None else ""
        base = base if base else "COLUMN"
        count = seen.get(base, 0) + 1
        seen[base] = count
        unique.append(base if count == 1 else f"{base}_{count}")
    return unique


def standardize_column_name(name: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(name).strip().upper()).strip("_")


def pick_column(columns: list[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def _normalize_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (bytes, bytearray)):
        try:
            return value.decode("utf-8")
        except Exception:
            return value.decode("latin-1", errors="ignore")
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _workbook_to_polars(workbook: dict) -> pl.DataFrame:
    frames = []
    for sheet_name, pdf in workbook.items():
        if pdf is None or pdf.empty:
            continue

        for col in pdf.columns:
            if pdf[col].dtype == "object":
                pdf[col] = pdf[col].map(_normalize_excel_value)

        pdf.columns = make_unique_columns(pdf.columns)
        pdf["_SOURCE_SHEET"] = str(sheet_name)
        frames.append(pl.from_pandas(pdf))

    if not frames:
        raise ValueError("Endorsement workbook has no readable rows.")

    return pl.concat(frames, how="diagonal_relaxed")


def _worksheet_to_polars(ws) -> pl.DataFrame | None:
    preview_rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
    if not preview_rows:
        return None

    def row_score(row) -> int:
        tokens = {
            str(cell).strip().upper()
            for cell in row
            if cell is not None and str(cell).strip() != ""
        }
        markers = {
            "PLACEMENT",
            "ACCOUNT NUMBER",
            "RECEIVED DATE",
            "SUB CAMPAIGN",
            "ENDO DATE",
            "PAYMENT AMOUNT",
        }
        return len(tokens & markers)

    header_idx = 0
    best_score = -1
    for idx, row in enumerate(preview_rows):
        score = row_score(row)
        if score > best_score:
            best_score = score
            header_idx = idx

    header = preview_rows[header_idx]
    if header is None:
        return None

    header = make_unique_columns(list(header))
    width = len(header)

    data_rows = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        normalized = [_normalize_excel_value(value) for value in list(row[:width])]
        if len(normalized) < width:
            normalized.extend([None] * (width - len(normalized)))
        if any(value is not None and str(value).strip() != "" for value in normalized):
            data_rows.append(normalized)

    if not data_rows:
        return None

    data = {name: [row[idx] for row in data_rows] for idx, name in enumerate(header)}
    frame = pl.DataFrame(data, strict=False)
    return frame.with_columns(pl.lit(str(ws.title)).alias("_SOURCE_SHEET"))


def _read_decrypted_workbook(file_bytes: bytes, workbook_password: str) -> pl.DataFrame:
    decrypted = BytesIO()
    office = msoffcrypto.OfficeFile(BytesIO(file_bytes))
    office.load_key(password=workbook_password)
    office.decrypt(decrypted)
    decrypted.seek(0)

    workbook = load_workbook(decrypted, read_only=True, data_only=True)
    frames = []
    for ws in workbook.worksheets:
        frame = _worksheet_to_polars(ws)
        if frame is not None:
            frames.append(frame)

    if not frames:
        raise ValueError("Decrypted workbook has no readable sheets.")

    return pl.concat(frames, how="diagonal_relaxed")


def read_endorsement_file(file_obj, workbook_password: str | None = None):
    ext = Path(file_obj.name).suffix.lower()
    file_bytes = file_obj.getvalue()

    if ext == ".csv":
        pdf = pd.read_csv(BytesIO(file_bytes))
        pdf.columns = make_unique_columns(pdf.columns)
        pdf["_SOURCE_SHEET"] = "CSV"
        return pl.from_pandas(pdf)

    effective_password = workbook_password or DEFAULT_WORKBOOK_PASSWORD

    if effective_password:
        try:
            return _read_decrypted_workbook(file_bytes, effective_password)
        except Exception as exc:
            raise ValueError(f"Could not read workbook after decrypting with the provided password: {exc}")

    raise ValueError("Workbook password is required for this endorsement file.")


def to_text_expr(df: pl.DataFrame, column_name: str | None) -> pl.Expr:
    if column_name is None or column_name not in df.columns:
        return pl.lit("")
    return pl.col(column_name).cast(pl.Utf8).fill_null("")


def to_date_expr(df: pl.DataFrame, column_name: str | None) -> pl.Expr:
    if column_name is None or column_name not in df.columns:
        return pl.lit(None, dtype=pl.Date)
    txt = pl.col(column_name).cast(pl.Utf8).str.strip_chars().str.replace_all(r"\s+", " ")
    return pl.coalesce([
        pl.col(column_name).cast(pl.Date, strict=False),
        pl.col(column_name).cast(pl.Datetime, strict=False).dt.date(),
        txt.str.strptime(pl.Date, "%m/%d/%Y", strict=False),
        txt.str.strptime(pl.Date, "%Y-%m-%d", strict=False),
        txt.str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).dt.date(),
        txt.str.strptime(pl.Datetime, "%m/%d/%Y %H:%M:%S", strict=False).dt.date(),
        txt.str.strptime(pl.Date, "%d/%m/%Y", strict=False),
    ])


def prepare_endorsement(endo_df: pl.DataFrame) -> pl.DataFrame:
    if endo_df.is_empty():
        raise ValueError("Endorsement data is empty.")

    std_names = [standardize_column_name(col) for col in endo_df.columns]
    rename_map = {}
    seen = {}
    for old_col, std_col in zip(endo_df.columns, std_names):
        count = seen.get(std_col, 0) + 1
        seen[std_col] = count
        rename_map[old_col] = std_col if count == 1 else f"{std_col}_{count}"

    df = endo_df.rename(rename_map)

    account_col = pick_column(df.columns, ["ACCOUNT_NO", "ACCOUNT_NUM", "ACCOUNT", "ACCOUNT_NUMBER"])
    agency_col = pick_column(df.columns, ["AGENCY", "AGENCY_NAME", "ENDORSEMENT_AGENCY", "PLACEMENT"])
    received_date_col = pick_column(df.columns, ["RECEIVED_DATE", "DATE_OF_ASSIGNMENT", "AS_OF"])
    pulled_out_date_col = pick_column(df.columns, ["PULLED_OUT_DATE", "PULLED OUT DATE", "PULLED_OUT", "PULLED DATE"])
    source_sheet_col = pick_column(df.columns, ["SOURCE_SHEET", "_SOURCE_SHEET"])
    ob_col = pick_column(df.columns, ["OB", "OUTSTANDING_BALANCE", "BALANCE", "OSB"])
    campaign_col = pick_column(df.columns, ["CAMPAIGN", "CAMPAIGN_NAME"])
    sub_campaign_col = pick_column(df.columns, ["SUB_CAMPAIGN", "SUBCAMPAIGN", "CAMPAIGN"])
    payment_date_col = pick_column(df.columns, ["PAYMENT_DATE", "PAYMENT DATE", "DATE"])
    endo_date_col = pick_column(df.columns, ["ENDO_DATE", "END_DATE", "PAYMENT_DATE", "DATE"])
    payment_amount_col = pick_column(
        df.columns,
        ["PAYMENT_AMOUNT", "AMOUNT", "COLLECTED_AMOUNT", "COLLECTED", "PAID_AMOUNT"],
    )

    if account_col is None:
        raise ValueError("Missing account column. Expected ACCOUNT/ACCOUNT_NO/ACCOUNT_NUMBER.")

    return df.with_columns(
        [
            to_text_expr(df, account_col).alias("ACCOUNT_KEY"),
            to_text_expr(df, agency_col).alias("AGENCY_STD"),
            to_text_expr(df, source_sheet_col).str.to_uppercase().str.strip_chars().alias("SOURCE_SHEET_STD"),
            to_date_expr(df, received_date_col).alias("RECEIVED_DATE_STD"),
            to_date_expr(df, pulled_out_date_col).alias("PULLED_OUT_DATE_STD"),
            (pl.col(ob_col).cast(pl.Float64, strict=False).fill_null(0.0) if ob_col else pl.lit(0.0)).alias("OB_STD"),
            to_text_expr(df, campaign_col).str.to_uppercase().str.strip_chars().alias("CAMPAIGN_STD"),
            to_text_expr(df, sub_campaign_col).str.to_uppercase().str.strip_chars().alias("SUB_CAMPAIGN_STD"),
            to_date_expr(df, payment_date_col).alias("PAYMENT_DATE_STD"),
            to_date_expr(df, endo_date_col).alias("ENDO_DATE_STD"),
            (pl.col(payment_amount_col).cast(pl.Float64, strict=False).fill_null(0.0) if payment_amount_col else pl.lit(0.0)).alias("PAYMENT_AMOUNT_STD"),
        ]
    )


def calculate_endorsement_metrics(
    df: pl.DataFrame,
    source_sheet: str,
    month: int,
    year: int,
    use_endorsement_rules: bool,
) -> pl.DataFrame:
    source_value = source_sheet.strip().upper()
    if use_endorsement_rules:
        filtered = df.filter(
            (pl.col("SOURCE_SHEET_STD") == source_value)
            & (pl.col("RECEIVED_DATE_STD").is_not_null())
            & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
            & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        )
        metric_name = "#_OF_ENDORSED_ACCOUNTS_HANDLED"
        bucket_label = f"{source_value} / RULES"
    else:
        filtered = df
        metric_name = "#_OF_ENDORSED_ACCOUNTS_HANDLED_DEFAULT"
        bucket_label = "DEFAULT FORMAT"

    unique_accounts = filtered.select(pl.col("ACCOUNT_KEY").n_unique()).item() if filtered.height > 0 else 0

    return pl.DataFrame(
        [
            {
                "METRIC": metric_name,
                "SOURCE_SHEET": bucket_label,
                "MONTH": month,
                "YEAR": year,
                "VALUE": int(unique_accounts),
            }
        ]
    )


def export_summary_excel(summary_df: pl.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_pandas().to_excel(writer, index=False, sheet_name="Summary")
    output.seek(0)
    return output.getvalue()


def month_name(month: int) -> str:
    return date(2000, int(month), 1).strftime("%B")


def previous_month_pair(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return 12, year - 1
    return month - 1, year


def extract_mmddyyyy_from_filename(file_name: str) -> date | None:
    match = re.search(r"(?<!\d)(\d{8})(?!\d)", file_name or "")
    if not match:
        return None

    raw_token = match.group(1)
    try:
        return datetime.strptime(raw_token, "%m%d%Y").date()
    except ValueError:
        return None


def set_derived_date_state(source_name: str, parsed_date: date) -> None:
    current_month = int(parsed_date.month)
    current_year = int(parsed_date.year)
    previous_month, previous_year = previous_month_pair(current_year, current_month)

    st.session_state[DERIVED_DATE_STATE_KEY] = {
        "source_file": source_name,
        "current_month": current_month,
        "current_year": current_year,
        "previous_month": int(previous_month),
        "previous_year": int(previous_year),
    }


def calculate_agency_input_table_121_150(
    prepared_df: pl.DataFrame,
    month: int,
    year: int,
) -> pl.DataFrame:
    target_placement = "MAYA CREDIT 121 - 150 DPD"
    normalized_target_placement = re.sub(r"\s+", " ", target_placement.strip().upper())
    placement_expr = pl.col("AGENCY_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    sheet_expr = pl.col("SOURCE_SHEET_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.strip_chars()
    sub_campaign_expr = pl.col("SUB_CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()

    active_filtered = prepared_df.filter(
        (sheet_expr == "ACTIVE")
        & (placement_expr == normalized_target_placement)
        & pl.col("RECEIVED_DATE_STD").is_not_null()
        & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
    )
    pout_filtered = prepared_df.filter(
        (sheet_expr == "POUT")
        & (placement_expr == normalized_target_placement)
        & pl.col("RECEIVED_DATE_STD").is_not_null()
        & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
    )
    endorsed_filtered = pl.concat([active_filtered, pout_filtered], how="diagonal_relaxed")
    endorsed_count = (
        endorsed_filtered.filter(pl.col("ACCOUNT_KEY").str.strip_chars() != "")
        .select(pl.col("ACCOUNT_KEY").n_unique())
        .item()
        if endorsed_filtered.height > 0
        else 0
    )

    pulled_out_filtered = pout_filtered
    pulled_out_count = pulled_out_filtered.height

    osb_filtered = endorsed_filtered.unique(subset=["ACCOUNT_KEY"], keep="first")
    osb_endorsed = osb_filtered.select(pl.col("OB_STD").sum()).item() if osb_filtered.height > 0 else 0.0

    payments_121 = prepared_df.filter(
        (sheet_expr == "PAYMENTS")
        & (sub_campaign_expr == normalized_target_placement)
        & pl.col("ENDO_DATE_STD").is_not_null()
        & (pl.col("ENDO_DATE_STD").dt.year() == year)
        & (pl.col("ENDO_DATE_STD").dt.month() == month)
    )
    selectives_121 = prepared_df.filter(
        (sheet_expr == "SELECTIVES")
        & (sub_campaign_expr == normalized_target_placement)
        & pl.col("ENDO_DATE_STD").is_not_null()
        & (pl.col("ENDO_DATE_STD").dt.year() == year)
        & (pl.col("ENDO_DATE_STD").dt.month() == month)
    )

    payments_sum = payments_121.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if payments_121.height > 0 else 0.0
    selectives_sum = selectives_121.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if selectives_121.height > 0 else 0.0
    collected_amount = float((payments_sum or 0.0) + (selectives_sum or 0.0))

    st.caption(
        "Collected breakdown - PAYMENTS 121-150 "
        f"({month_name(month)}): {payments_sum or 0.0:,.2f} | SELECTIVES 121-150 "
        f"({month_name(month)}): {selectives_sum or 0.0:,.2f}"
    )

    return pl.DataFrame(
        [
            {
                "Bucket": "121-150 DPD",
                "Agency": "SP MADRID",
                "# of Endorsed Accounts Handled": int(endorsed_count),
                "# of Accounts Pulled Out": int(pulled_out_count),
                "OSB Endorsed (₱)": float(osb_endorsed or 0.0),
                "Collected (₱)": collected_amount,
            }
        ]
    )


def calculate_agency_input_table_181_above(
    prepared_df: pl.DataFrame,
    collected_month: int,
    collected_year: int,
) -> pl.DataFrame:
    target_placement = "MAYA CREDIT 181 DPD & UP"
    normalized_target_placement = re.sub(r"\s+", " ", target_placement.strip().upper())
    placement_expr = pl.col("AGENCY_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    sheet_expr = pl.col("SOURCE_SHEET_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.strip_chars()
    sub_campaign_expr = pl.col("SUB_CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    campaign_expr = pl.col("CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()

    active_filtered = prepared_df.filter(
        (sheet_expr == "ACTIVE")
        & (placement_expr == normalized_target_placement)
    )

    pout_endorsed_filtered = prepared_df.filter(
        (sheet_expr == "POUT")
        & (placement_expr == normalized_target_placement)
        & pl.col("PULLED_OUT_DATE_STD").is_not_null()
        & (pl.col("PULLED_OUT_DATE_STD").dt.year() == collected_year)
        & (pl.col("PULLED_OUT_DATE_STD").dt.month() == collected_month)
    )
    endorsed_filtered = pl.concat([active_filtered, pout_endorsed_filtered], how="diagonal_relaxed")
    endorsed_count = (
        endorsed_filtered.filter(pl.col("ACCOUNT_KEY").str.strip_chars() != "")
        .select(pl.col("ACCOUNT_KEY").n_unique())
        .item()
        if endorsed_filtered.height > 0
        else 0
    )

    pulled_out_count = pout_endorsed_filtered.height

    osb_filtered = endorsed_filtered.unique(subset=["ACCOUNT_KEY"], keep="first")
    osb_endorsed = osb_filtered.select(pl.col("OB_STD").sum()).item() if osb_filtered.height > 0 else 0.0

    payments_181 = prepared_df.filter(
        sheet_expr.is_in(["PAYMENT", "PAYMENTS"])
        & (sub_campaign_expr == normalized_target_placement)
        & (campaign_expr == "MAYA CREDIT")
        & pl.col("PAYMENT_DATE_STD").is_not_null()
        & (pl.col("PAYMENT_DATE_STD").dt.month() == collected_month)
        & (pl.col("PAYMENT_DATE_STD").dt.year() == collected_year)
    )

    payments_sum = payments_181.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if payments_181.height > 0 else 0.0
    collected_amount = float(payments_sum or 0.0)

    st.caption(
        f"Collected breakdown - PAYMENTS 181 {month_name(collected_month)} only: "
        f"{payments_sum or 0.0:,.2f}"
    )

    return pl.DataFrame(
        [
            {
                "Bucket": "181+ DPD",
                "Agency": "SP MADRID",
                "# of Endorsed Accounts Handled": int(endorsed_count),
                "# of Accounts Pulled Out": int(pulled_out_count),
                "OSB Endorsed (₱)": float(osb_endorsed or 0.0),
                "Collected (₱)": collected_amount,
            }
        ]
    )


def render_masterfile_tab() -> None:
    st.subheader("Masterfile")
    st.caption("Endorsement-only processing")

    with st.expander("Options", expanded=True):
        use_server_file = st.checkbox("Use server file by name", value=True, key="masterfile_use_server")
        server_file_name = ""
        if use_server_file:
            st.caption(f"Server source: {DEFAULT_SERVER_MASTERFILE_DIR}")
            server_file_name = st.text_input(
                "Server file name",
                value=DEFAULT_ENDORSEMENT_FILE,
                help="Paste filename, relative path (e.g. MARCH 2026\\file.xlsx), or full UNC file path.",
                key="masterfile_server_file_name",
            )
        workbook_password = st.text_input(
            "Workbook password",
            value=DEFAULT_WORKBOOK_PASSWORD,
            type="password",
            key="masterfile_workbook_password",
        )

    st.write("Upload endorsement files (optional if local folder mode is enabled).")
    uploaded_files = st.file_uploader(
        "Upload Endorsement File(s)",
        type=["xlsx", "xls", "xlsb", "csv"],
        accept_multiple_files=True,
        key="masterfile_uploader",
    )

    if st.button("Submit", type="secondary", use_container_width=True, key="masterfile_submit"):
        progress = st.progress(0)
        status = st.empty()

        def set_progress(percent: int, message: str):
            progress.progress(percent)
            status.caption(f"{percent}% - {message}")

        try:
            set_progress(10, "Collecting files")
            selected_file = None

            if use_server_file:
                resolved_server_file, resolve_error = resolve_server_endorsement_file(server_file_name)
                if resolve_error:
                    set_progress(100, "Stopped")
                    st.error(resolve_error)
                    return

                selected_file = LocalInputFile(resolved_server_file)
                st.caption(f"Using server file: {selected_file.path}")

            if selected_file is None and uploaded_files:
                selected_file = uploaded_files[0]
                st.caption(f"Using uploaded file: {selected_file.name}")

            if selected_file is None:
                set_progress(100, "Stopped")
                st.error("No endorsement file selected. Upload a file or enable server file mode with a valid file name.")
                return

            derived_date = extract_mmddyyyy_from_filename(selected_file.name)
            if derived_date is None:
                set_progress(100, "Stopped")
                st.error(
                    "Could not find a valid mmddyyyy date in the selected file name. "
                    "Rename the file (example: MAYA ENDORSEMENT 04082026.xlsx) and try again."
                )
                return

            set_derived_date_state(selected_file.name, derived_date)
            date_state = st.session_state.get(DERIVED_DATE_STATE_KEY, {})
            current_month = int(date_state.get("current_month"))
            current_year = int(date_state.get("current_year"))
            previous_month = int(date_state.get("previous_month"))
            previous_year = int(date_state.get("previous_year"))

            set_progress(35, "Reading endorsement workbook")
            raw_df = read_endorsement_file(
                selected_file,
                workbook_password=workbook_password.strip() or None,
            )

            set_progress(65, "Preparing endorsement columns")
            prepared_df = prepare_endorsement(raw_df)

            set_progress(85, "Calculating metrics")
            st.subheader("Agency Input Calculation Table")
            st.caption(
                f"Month flow: 121 = {month_name(int(previous_month))}, 181 = {month_name(int(current_month))}."
            )
            calc_table_121_150 = calculate_agency_input_table_121_150(
                prepared_df,
                month=int(previous_month),
                year=int(previous_year),
            )
            calc_table_181_above = calculate_agency_input_table_181_above(
                prepared_df,
                collected_month=int(current_month),
                collected_year=int(current_year),
            )
            combined_table = pl.concat([calc_table_121_150, calc_table_181_above], how="diagonal_relaxed")
            calc_pd = combined_table.to_pandas()
            st.dataframe(
                calc_pd.style.format(
                    {
                        "# of Endorsed Accounts Handled": "{:,.0f}",
                        "# of Accounts Pulled Out": "{:,.0f}",
                        "OSB Endorsed (₱)": "{:,.2f}",
                        "Collected (₱)": "{:,.2f}",
                    }
                ),
                use_container_width=True,
            )

            st.subheader("Excel Preview")
            st.caption(
                f"Previewing {selected_file.name} with 121 set to {month_name(int(previous_month))} and 181 set to {month_name(int(current_month))}."
            )
            st.dataframe(prepared_df.to_pandas().head(50), use_container_width=True)

            set_progress(100, "Completed")
            st.success("Endorsement metrics generated successfully.")
        except Exception as exc:
            set_progress(100, "Failed")
            st.error(f"Failed to process endorsement file: {exc}")


def _keep_drr_column(col_name: str) -> bool:
    normalized = col_name.lower().replace(" ", "_")
    return normalized in USED_COLUMNS


def _duration_seconds(series: pd.Series) -> pd.Series:
    parts = series.astype(str).str.split(":", expand=True)
    if parts.shape[1] >= 3:
        return (
            pd.to_numeric(parts[0], errors="coerce").fillna(0) * 3600
            + pd.to_numeric(parts[1], errors="coerce").fillna(0) * 60
            + pd.to_numeric(parts[2], errors="coerce").fillna(0)
        )
    return pd.Series(0, index=series.index)


def _extract_drr_category_rows(df_norm: pd.DataFrame, raw_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    result = {name: pd.DataFrame() for name in DRR_CATEGORY_STATUS}
    if STATUS_COLUMN not in df_norm.columns:
        return result

    working = df_norm.copy()
    if "account_no." in working.columns:
        acct = working["account_no."]
        numeric = pd.to_numeric(acct, errors="coerce")
        acct_normalized = numeric.where(numeric.isna(), numeric.astype("int64").astype(str))
        acct_normalized = acct_normalized.where(acct_normalized.notna(), acct.astype(str).str.strip())
        acct_normalized = acct_normalized.where(acct.notna(), None)
        working["account_normalized"] = acct_normalized
        working = working[working["account_normalized"].fillna("").str.startswith("6")]

    if "call_duration" in working.columns:
        connected_mask = _duration_seconds(working["call_duration"]) > 0
        connected = working[connected_mask]
    else:
        connected = pd.DataFrame(columns=working.columns)

    if connected.empty:
        return result

    connected_status = connected[STATUS_COLUMN].fillna("").str.upper()
    for sheet_name, statuses in DRR_CATEGORY_STATUS.items():
        matched_idx = connected[connected_status.isin(statuses)].index
        if len(matched_idx) > 0:
            result[sheet_name] = raw_df.loc[matched_idx].copy()

    return result

# ── DRR Tab Logic ──────────────────────────────────────────

def compute_metrics(df: pd.DataFrame) -> Dict:
    status = df[STATUS_COLUMN].fillna("").str.upper()

    has_account = "account_no." in df.columns
    if has_account:
        acct = df["account_no."]
        numeric = pd.to_numeric(acct, errors="coerce")
        acct_normalized = numeric.where(numeric.isna(), numeric.astype("int64").astype(str))
        acct_normalized = acct_normalized.where(acct_normalized.notna(), acct.astype(str).str.strip())
        acct_normalized = acct_normalized.where(acct.notna(), None)
        df = df.copy()
        df["account_normalized"] = acct_normalized
        df = df[df["account_normalized"].fillna("").str.startswith("6")]
        status = df[STATUS_COLUMN].fillna("").str.upper()

    if "call_duration" in df.columns:
        s     = df["call_duration"].astype(str)
        split = s.str.split(":", expand=True)
        if split.shape[1] >= 3:
            dur = (
                pd.to_numeric(split[0], errors="coerce").fillna(0) * 3600 +
                pd.to_numeric(split[1], errors="coerce").fillna(0) * 60 +
                pd.to_numeric(split[2], errors="coerce").fillna(0)
            )
        else:
            dur = pd.Series(0, index=df.index)
        connected        = df[dur > 0]
        connected_status = connected[STATUS_COLUMN].fillna("").str.upper()
    else:
        connected        = pd.DataFrame()
        connected_status = pd.Series(dtype=str)

    has_remark_type = "remark_type" in df.columns
    has_remark      = "remark" in df.columns

    # Single-pass status flag columns
    has_email = status.str.contains("EMAIL", na=False)
    has_sms   = status.str.contains("SMS",   na=False)
    has_untc  = status.str.contains("UNTC",  na=False)
    has_viber = status.str.contains("VIBER", na=False)
    exclude_filter = (~has_email) | (~has_sms) | (~has_untc) | (~has_viber)

    if has_remark_type:
        rt = df["remark_type"].fillna("").str.strip()
        manual_dials = int((exclude_filter & (rt == "Outgoing")).sum())
        dials        = int((exclude_filter & (rt != "Outgoing")).sum())
    else:
        manual_dials = 0
        dials        = int(exclude_filter.sum())

    account_lists = {}
    connected_unique = 0
    rpc_under_nego = 0
    rpc_niop = 0
    third_party = 0
    dispute = 0
    email_sent = 0
    email_responsive = 0

    if has_account:
        account_lists = {
            "connected_accounts":        connected["account_normalized"].dropna().tolist(),
            "rpc_under_nego_accounts":   connected[connected_status.isin(["POSITIVE CONTACT - CALLBACK", "POSITIVE CONTACT - UNDERNEGO"])]["account_normalized"].dropna().tolist(),
            "rpc_niop_accounts":         connected[connected_status.isin(["POSITIVE CONTACT - RPC REFUSE TO PAY", "POSITIVE CONTACT - DISPUTE"])]["account_normalized"].dropna().tolist(),
            "third_party_accounts":      connected[connected_status == "POSITIVE - 3RD PARTY CONTACTED"]["account_normalized"].dropna().tolist(),
            "dispute_accounts":          connected[connected_status == "POSITIVE CONTACT - DISPUTE"]["account_normalized"].dropna().tolist(),
            "email_sent_accounts":       df[status.str.contains("SENT EMAIL",       na=False)]["account_normalized"].dropna().tolist(),
            "email_responsive_accounts": df[status.str.contains("EMAIL RESPONSIVE", na=False)]["account_normalized"].dropna().tolist(),
        }

        connected_unique = len(set(account_lists.get("connected_accounts", [])))
        rpc_under_nego = len(set(account_lists.get("rpc_under_nego_accounts", [])))
        rpc_niop = len(set(account_lists.get("rpc_niop_accounts", [])))
        third_party = len(set(account_lists.get("third_party_accounts", [])))
        dispute = len(set(account_lists.get("dispute_accounts", [])))
        email_sent = len(set(account_lists.get("email_sent_accounts", [])))
        email_responsive = len(set(account_lists.get("email_responsive_accounts", [])))

    return {
        "dials":         dials,
        "manual_dials":  manual_dials,
        "connected":     len(connected),
        "connected_unique": connected_unique,
        "rpc_under_nego": rpc_under_nego,
        "rpc_niop": rpc_niop,
        "third_party": third_party,
        "dispute": dispute,
        "email_sent": email_sent,
        "email_responsive": email_responsive,
        "vb_count":      int(df["remark"].fillna("").str.contains("broadcast", case=False, na=False).sum()) if has_remark else 0,
        "vb_connected":  int(status.isin(["PU", "PM"]).sum()),
        "account_lists": account_lists,
    }


def process_single_file(file, cycle_1: str, cycle_2: str) -> Tuple[str, Dict, int, int, int, Dict[str, pd.DataFrame]]:
    try:
        # calamine is a Rust-based reader — significantly faster than openpyxl
        raw = pd.read_excel(file, engine="calamine", dtype=str)
        raw_norm = raw.copy()
        raw_norm.columns = raw_norm.columns.str.lower().str.replace(" ", "_", regex=False)
        total_rows = len(raw)

        keep = [c for c in raw_norm.columns if c in USED_COLUMNS]
        df = raw_norm[keep]

        if STATUS_COLUMN not in df.columns:
            return file.name, {}, total_rows, 0, 0, {name: pd.DataFrame() for name in DRR_CATEGORY_STATUS}

        has_cycle     = CYCLE_COLUMN in df.columns
        cycle_results = {}
        category_frames = {name: [] for name in DRR_CATEGORY_STATUS}
        c1_rows = c2_rows = 0

        if has_cycle:
            cycle_col = df[CYCLE_COLUMN].astype(str).str.strip()

        if cycle_1:
            df_c1   = df[cycle_col == cycle_1] if has_cycle else df
            c1_rows = len(df_c1)
            cycle_results[cycle_1] = compute_metrics(df_c1)
            extracted = _extract_drr_category_rows(df_c1, raw)
            for sheet_name, frame in extracted.items():
                if not frame.empty:
                    category_frames[sheet_name].append(frame)

        if cycle_2:
            df_c2   = df[cycle_col == cycle_2] if has_cycle else df
            c2_rows = len(df_c2)
            cycle_results[cycle_2] = compute_metrics(df_c2)
            extracted = _extract_drr_category_rows(df_c2, raw)
            for sheet_name, frame in extracted.items():
                if not frame.empty:
                    category_frames[sheet_name].append(frame)

        if not cycle_1 and not cycle_2:
            c1_rows = total_rows
            cycle_results["All Data"] = compute_metrics(df)
            extracted = _extract_drr_category_rows(df, raw)
            for sheet_name, frame in extracted.items():
                if not frame.empty:
                    category_frames[sheet_name].append(frame)

        category_rows = {
            sheet_name: (pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=raw.columns))
            for sheet_name, frames in category_frames.items()
        }

        return file.name, cycle_results, total_rows, c1_rows, c2_rows, category_rows

    except Exception as e:
        st.error(f"Error processing {file.name}: {e}")
        return file.name, {}, 0, 0, 0, {name: pd.DataFrame() for name in DRR_CATEGORY_STATUS}


def aggregate_results(file_results: Dict) -> Dict:
    overall    = {}
    all_cycles = set()
    for d in file_results.values():
        all_cycles.update(d["cycle_status_counts"].keys())

    for cycle in all_cycles:
        totals            = defaultdict(int)
        account_aggregates = defaultdict(list)

        for d in file_results.values():
            if cycle in d["cycle_status_counts"]:
                metrics = d["cycle_status_counts"][cycle]
                for m in ["dials", "manual_dials", "connected", "vb_count", "vb_connected"]:
                    totals[m] += metrics.get(m, 0)
                for k, v in metrics.get("account_lists", {}).items():
                    account_aggregates[k].extend(v)

        totals["connected_unique"] = len(set(account_aggregates.get("connected_accounts", [])))
        totals["rpc_under_nego"]   = len(set(account_aggregates.get("rpc_under_nego_accounts", [])))
        totals["rpc_niop"]         = len(set(account_aggregates.get("rpc_niop_accounts", [])))
        totals["third_party"]      = len(set(account_aggregates.get("third_party_accounts", [])))
        totals["dispute"]          = len(set(account_aggregates.get("dispute_accounts", [])))
        totals["email_sent"]       = len(set(account_aggregates.get("email_sent_accounts", [])))
        totals["email_responsive"] = len(set(account_aggregates.get("email_responsive_accounts", [])))

        overall[cycle] = dict(totals)

    return overall


def build_excel(file_results: Dict, overall_counts: Dict, cycle_1: str, cycle_2: str) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Overall Summary
        overall_data = {"Metric": list(METRIC_LABELS.values())}
        for cycle, metrics in overall_counts.items():
            overall_data[cycle] = [metrics.get(k, 0) for k in METRIC_LABELS]
        pd.DataFrame(overall_data).to_excel(writer, sheet_name="Overall Summary", index=False)

        # Per-File Summary — one row per file+cycle
        per_file_rows = []
        for fname, d in file_results.items():
            for cycle, metrics in d["cycle_status_counts"].items():
                row = {"File": fname, "Cycle": cycle}
                for k, label in METRIC_LABELS.items():
                    row[label] = metrics.get(k, 0)
                per_file_rows.append(row)
        pd.DataFrame(per_file_rows).to_excel(writer, sheet_name="Per-File Summary", index=False)

        # File Statistics
        stats = [
            {
                "File":            fname,
                "Total Rows":      d["total_rows"],
                f"{cycle_1} Rows": d["cycle_1_filtered"],
                f"{cycle_2} Rows": d["cycle_2_filtered"],
            }
            for fname, d in file_results.items()
        ]
        pd.DataFrame(stats).to_excel(writer, sheet_name="File Statistics", index=False)

        for sheet_name in DRR_CATEGORY_STATUS:
            category_frames = [
                d.get("category_rows", {}).get(sheet_name)
                for d in file_results.values()
                if not d.get("category_rows", {}).get(sheet_name, pd.DataFrame()).empty
            ]
            if category_frames:
                pd.concat(category_frames, ignore_index=True).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)

        for ws in writer.sheets.values():
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output.seek(0)
    return output.getvalue()


# ── Session State ─────────────────────────────────────────────────────────────
for key in ["drr_results", "drr_overall", "drr_excel", "drr_cycle_1", "drr_cycle_2"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_masterfile, tab_drr = st.tabs(["Masterfile", "DRR"])

with tab_drr:
    col1, col2 = st.columns(2)
    cycle_1_input = col1.text_input("Cycle 1", value="MC 121-150 DPD")
    cycle_2_input = col2.text_input("Cycle 2", value="MC 181DPD UP")

    uploaded_files = st.file_uploader(
        "Upload Excel File/s", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True
    )

    if st.button("Process", use_container_width=True, disabled=not uploaded_files):
        file_results = {}
        total_files  = len(uploaded_files)
        with st.status("Processing files...", expanded=True) as status:
            progress = st.progress(0, text="Starting...")
            for i, f in enumerate(uploaded_files):
                progress.progress(i / total_files, text=f"Processing {f.name} ({i + 1}/{total_files})...")
                fname, cycle_counts, total, c1, c2, category_rows = process_single_file(f, cycle_1_input, cycle_2_input)
                file_results[fname] = {
                    "cycle_status_counts": cycle_counts,
                    "total_rows":          total,
                    "cycle_1_filtered":    c1,
                    "cycle_2_filtered":    c2,
                    "category_rows":       category_rows,
                }
            progress.progress(1.0, text="Finalizing...")
            status.update(label="Done", state="complete", expanded=False)

        overall = aggregate_results(file_results)
        st.session_state.drr_results  = file_results
        st.session_state.drr_overall  = overall
        st.session_state.drr_excel    = build_excel(file_results, overall, cycle_1_input, cycle_2_input)
        st.session_state.drr_cycle_1  = cycle_1_input
        st.session_state.drr_cycle_2  = cycle_2_input

    if st.session_state.drr_overall:
        st.divider()

        # Summary metrics per cycle
        for cycle, metrics in st.session_state.drr_overall.items():
            st.subheader(cycle)
            rows = [{"Metric": METRIC_LABELS[k], "Value": f"{metrics.get(k, 0):,}"} for k in METRIC_LABELS]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # File statistics
        st.subheader("File Statistics")
        c1 = st.session_state.drr_cycle_1
        c2 = st.session_state.drr_cycle_2
        stats_rows = [
            {
                "File":            fname,
                "Total Rows":      d["total_rows"],
                f"{c1} Rows":      d["cycle_1_filtered"],
                f"{c2} Rows":      d["cycle_2_filtered"],
            }
            for fname, d in st.session_state.drr_results.items()
        ]
        st.dataframe(pd.DataFrame(stats_rows), use_container_width=True, hide_index=True)

        st.download_button(
            label="Download Summary (.xlsx)",
            data=st.session_state.drr_excel,
            file_name="drr_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

with tab_masterfile:
    render_masterfile_tab()