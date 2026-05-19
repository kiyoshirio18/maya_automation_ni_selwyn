import os
import json
import re
import zipfile
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PLACEMENT_VALUES = [
    "Maya Credit 181 DPD & UP",
    "Maya Credit 121 - 150 DPD",
    "Maya Negosyo Advance 121 - 150 DPD",
    "Maya Negosyo Advance 181 DPD & UP",
]
PLACEMENT_METRIC_LABELS = {
    "Maya Credit 181 DPD & UP": "Credit 181+",
    "Maya Credit 121 - 150 DPD": "Credit 121-150",
    "Maya Negosyo Advance 121 - 150 DPD": "Negosyo 121-150",
    "Maya Negosyo Advance 181 DPD & UP": "Negosyo 181+",
}

SOURCE_SHEET = "ACTIVE"
OUTPUT_SHEET_NAME = "Digital Result"
DIGITAL_DISPO_SHEET_NAME = "Digital Dispo"
DIGITAL_DISPO_COLUMN_NAME = "Digital Dispo"
DIGITAL_DISPO_HEADER_FILL = "FFD9E1F2"
DIGITAL_DISPO_VALUES = [
    "EMAIL - FAILED",
    "EMAIL - RECEIVED - PAYMENT PROCESS ERROR - CAN'T PAY",
    "EMAIL - RECEIVED - PROMISED TO PAY",
    "EMAIL - RECEIVED - REFUSED TO PAY",
    "EMAIL - RECEIVED - REQUEST FOR CALLBACK",
    "EMAIL - RECEIVED - UNAWARE OF LOAN - WON'T PAY",
    "EMAIL - SENT",
    "OTHER NON-VOICE - FAILED",
    "OTHER NON-VOICE - RECEIVED - PAYMENT PROCESS ERROR - CAN'T PAY",
    "OTHER NON-VOICE - RECEIVED - PROMISED TO PAY",
    "OTHER NON-VOICE - RECEIVED - REFUSED TO PAY",
    "OTHER NON-VOICE - RECEIVED - REQUEST FOR CALLBACK",
    "OTHER NON-VOICE - RECEIVED - UNAWARE OF LOAN - WON'T PAY",
    "OTHER NON-VOICE - SENT",
    "SCH - NON-VOICE - Claims Paid; Missing Record",
    "SCH - NON-VOICE - Complaints",
    "SCH - NON-VOICE - Deceased",
    "SCH - NON-VOICE - Dispute",
    "SCH - NON-VOICE - Hardship",
    "SCH - NON-VOICE - Incarcerated RPC",
    "SCH - NON-VOICE - Insolvency/Bankruptcy",
    "SCH - NON-VOICE - Possible Fraud",
    "SCH - NON-VOICE - Vulnerable Customer",
    "SCH - NON-VOICE - Wrong Number",
    "SMS - FAILED",
    "SMS - RECEIVED - PAYMENT PROCESS ERROR - CAN'T PAY",
    "SMS - RECEIVED - PROMISED TO PAY",
    "SMS - RECEIVED - REFUSED TO PAY",
    "SMS - RECEIVED - REQUEST FOR CALLBACK",
    "SMS - RECEIVED - UNAWARE OF LOAN - WON'T PAY",
    "SMS - SENT",
    "VIBER - FAILED",
    "VIBER - RECEIVED - PAYMENT PROCESS ERROR - CAN'T PAY",
    "VIBER - RECEIVED - PROMISED TO PAY",
    "VIBER - RECEIVED - REFUSED TO PAY",
    "VIBER - RECEIVED - REQUEST FOR CALLBACK",
    "VIBER - RECEIVED - UNAWARE OF LOAN - WON'T PAY",
    "VIBER - SENT",
]
TEMPLATE_OUTPUT_COLUMNS = [
    "Agency",
    "Date of Blast",
    "Product",
    "Bucket",
    "Account Number",
    "Mobile Number",
    "SMS Dispo",
    "Email Address",
    "Email Dispo",
    "Email Respose",
    "Viber Dispo",
    "IVR Dispo",
    "# of Account Called Agency Hotline",
    "Using Forbearance template ? Y/N",
]
OUTPUT_COLUMN_MIN_WIDTH = 18
OUTPUT_COLUMN_MAX_WIDTH = 120
OUTPUT_COLUMN_PADDING = 4
OUTPUT_COLUMN_WIDTH_OVERRIDES = {
    "SMS Dispo": 28,
    "Email Dispo": 24,
    "IVR Dispo": 30,
    "# of Account Called Agency Hotline": 34,
    "Using Forbearance template ? Y/N": 34,
    "Account Number": 22,
    "Mobile Number": 20,
    "Email Address": 30,
}
REQUIRED_SOURCE_COLUMNS = [
    "PLACEMENT",
    "Product",
    "Bucket",
    "Account Number",
    "Mobile Number",
    "Email Address",
]
SOURCE_COLUMN_ALIASES = {
    "PLACEMENT": ["PLACEMENT"],
    "Product": ["Product", "PRODUCT", "PRODUCT_NAME"],
    "Bucket": ["Bucket", "BUCKET", "DPD BUCKET"],
    "Account Number": ["Account Number", "ACCOUNT NUMBER", "ACCOUNT_NUMBER"],
    "Mobile Number": ["Mobile Number", "MOBILE NUMBER", "MOBILE_NUMBER_1"],
    "Email Address": ["Email Address", "EMAIL ADDRESS", "EMAIL_ADDRESS"],
}

load_dotenv()


def _normalize_unc_path(path_text: str) -> str:
    normalized = (path_text or "").strip().replace("/", "\\")
    normalized = normalized.replace("\x07", r"\a")

    if normalized.startswith("\\") and not normalized.startswith("\\\\"):
        normalized = "\\" + normalized

    return normalized


DEFAULT_SERVER_MASTERFILE_DIR = _normalize_unc_path(
    os.getenv(
        "DEFAULT_SERVER_MASTERFILE_DIR",
        r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ENDO\MASTERFILE",
    )
)
DEFAULT_SERVER_MASTERFILE_FILE = os.getenv(
    "DEFAULT_SERVER_MASTERFILE_FILE",
    f"MAYA ENDORSEMENT {date.today():%m%d%Y}.xlsx",
)
MF_WB_PASS = os.getenv("MASTERFILE_WORKBOOK_PASSWORD", "Maya@2026")
DEFAULT_DAILY_DIGITAL_REPORT_DIR = _normalize_unc_path(
    os.getenv(
        "DEFAULT_DAILY_DIGITAL_REPORT_DIR",
        r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\DAILY DIGITAL REPORT",
    )
)
KPI_CACHE_FILE_PATH = Path(__file__).resolve().parents[1] / "resources" / "digital_daily_kpi_cache.json"


def _extract_output_date_from_filename(output_filename: str) -> datetime:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", output_filename)
    if not match:
        raise ValueError(
            f"Unable to extract output date from filename: {output_filename}"
        )
    return datetime.strptime(match.group(1), "%Y-%m-%d")


def build_output_destination_path(output_filename: str) -> Path:
    output_date = _extract_output_date_from_filename(output_filename)
    year_folder = f"{output_date.year}"
    month_year_folder = output_date.strftime("%B %Y").upper()
    return Path(DEFAULT_DAILY_DIGITAL_REPORT_DIR) / year_folder / month_year_folder / output_filename


def save_output_to_network(output_bytes: bytes, output_filename: str) -> Path:
    destination_path = build_output_destination_path(output_filename)
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    if destination_path.exists():
        raise FileExistsError(f"File already exists: {destination_path}")
    destination_path.write_bytes(output_bytes)
    return destination_path


def resolve_server_masterfile(server_input: str) -> tuple[Path | None, str | None]:
    master_dir = Path(DEFAULT_SERVER_MASTERFILE_DIR)
    if not master_dir.exists() or not master_dir.is_dir():
        return None, f"Server folder is not reachable: {master_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Server file mode is enabled but no file name/path was provided."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)

    if requested_path.exists() and requested_path.is_file():
        return requested_path, None

    candidate = master_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None

    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid server file name or relative path."

    matches = [
        p for p in master_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()
    ]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, (
            f"Multiple files named '{requested_name}' were found under MASTERFILE. "
            "Please paste a relative path, for example 'MARCH 2026\\file.xlsx'.\n"
            f"Matches:\n{preview}"
        )

    return None, f"Server file not found from input: {requested}"


def _normalize_column_name(column_name: object) -> str:
    text = str(column_name)
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()


def _normalize_placement_text(value: object) -> str:
    text = str(value)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _build_normalized_placement_series(dataframe: pd.DataFrame, column_name: str) -> pd.Series:
    return (
        dataframe[column_name]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )


def get_placement_counts(dataframe: pd.DataFrame) -> Dict[str, int]:
    counts = {placement: 0 for placement in PLACEMENT_VALUES}
    if "PLACEMENT" not in dataframe.columns:
        return counts

    normalized_series = _build_normalized_placement_series(dataframe, "PLACEMENT")
    non_null_mask = dataframe["PLACEMENT"].notna()

    for placement in PLACEMENT_VALUES:
        normalized_target = _normalize_placement_text(placement)
        counts[placement] = int((non_null_mask & normalized_series.eq(normalized_target)).sum())

    return counts


def _normalize_cached_counts(raw_counts: object) -> Dict[str, int]:
    normalized_counts = {placement: 0 for placement in PLACEMENT_VALUES}
    if not isinstance(raw_counts, dict):
        return normalized_counts

    for placement in PLACEMENT_VALUES:
        raw_value = raw_counts.get(placement, 0)
        try:
            normalized_counts[placement] = int(raw_value)
        except (TypeError, ValueError):
            normalized_counts[placement] = 0

    return normalized_counts


def _normalize_cache_entry(raw_entry: object) -> Dict[str, object]:
    normalized_entry: Dict[str, object] = {
        "counts": {placement: 0 for placement in PLACEMENT_VALUES},
        "total_rows": 0,
    }

    if not isinstance(raw_entry, dict):
        return normalized_entry

    if "counts" in raw_entry or "total_rows" in raw_entry:
        counts_raw = raw_entry.get("counts", {})
        total_rows_raw = raw_entry.get("total_rows")
    else:
        counts_raw = raw_entry
        total_rows_raw = None

    normalized_counts = _normalize_cached_counts(counts_raw)
    normalized_entry["counts"] = normalized_counts

    if total_rows_raw is None:
        normalized_entry["total_rows"] = sum(normalized_counts.values())
    else:
        try:
            normalized_entry["total_rows"] = int(total_rows_raw)
        except (TypeError, ValueError):
            normalized_entry["total_rows"] = sum(normalized_counts.values())

    return normalized_entry


def load_kpi_cache() -> Dict[str, Dict[str, object]]:
    if not KPI_CACHE_FILE_PATH.exists():
        return {}

    try:
        raw_text = KPI_CACHE_FILE_PATH.read_text(encoding="utf-8")
        payload = json.loads(raw_text)
    except Exception:
        return {}

    if not isinstance(payload, dict):
        return {}

    cache: Dict[str, Dict[str, object]] = {}
    for date_key, raw_entry in payload.items():
        if not isinstance(date_key, str):
            continue
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_key) is None:
            continue
        cache[date_key] = _normalize_cache_entry(raw_entry)

    return cache


def save_kpi_cache(cache: Dict[str, Dict[str, object]]) -> None:
    KPI_CACHE_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    KPI_CACHE_FILE_PATH.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def get_comparison_counts_from_cache(
    current_date: datetime,
    cache: Dict[str, Dict[str, object]],
) -> tuple[Optional[Dict[str, int]], Optional[int], Optional[datetime]]:
    current_key = current_date.strftime("%Y-%m-%d")
    yesterday_key = (current_date - timedelta(days=1)).strftime("%Y-%m-%d")

    if yesterday_key in cache:
        entry = cache[yesterday_key]
        return (
            entry.get("counts") if isinstance(entry.get("counts"), dict) else None,
            int(entry.get("total_rows", 0)),
            datetime.strptime(yesterday_key, "%Y-%m-%d"),
        )

    previous_keys = sorted(date_key for date_key in cache.keys() if date_key < current_key)
    if previous_keys:
        last_key = previous_keys[-1]
        entry = cache[last_key]
        return (
            entry.get("counts") if isinstance(entry.get("counts"), dict) else None,
            int(entry.get("total_rows", 0)),
            datetime.strptime(last_key, "%Y-%m-%d"),
        )

    return None, None, None


def upsert_kpi_cache(current_date: datetime, counts: Dict[str, int], total_rows: int) -> None:
    cache = load_kpi_cache()
    cache[current_date.strftime("%Y-%m-%d")] = {
        "counts": _normalize_cached_counts(counts),
        "total_rows": int(total_rows),
    }
    save_kpi_cache(cache)


def find_previous_masterfile_path(current_path: Path, current_date: datetime) -> Optional[Path]:
    if not current_path.exists():
        return None

    candidates: List[tuple[datetime, Path]] = []
    for pattern in ("*.xlsx", "*.xls", "*.xlsb"):
        for path in current_path.parent.glob(pattern):
            if path.resolve() == current_path.resolve():
                continue
            date_match = re.search(r"(\d{8})", path.stem)
            if not date_match:
                continue
            try:
                file_date = datetime.strptime(date_match.group(1), "%m%d%Y")
            except ValueError:
                continue
            if file_date < current_date:
                candidates.append((file_date, path))

    if not candidates:
        return None

    return max(candidates, key=lambda item: item[0])[1]


def _align_source_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    normalized_to_actual = {}
    for column in dataframe.columns:
        normalized = _normalize_column_name(column)
        if normalized not in normalized_to_actual:
            normalized_to_actual[normalized] = str(column)

    rename_map = {}
    missing_columns = []
    for required_column in REQUIRED_SOURCE_COLUMNS:
        candidate_headers = SOURCE_COLUMN_ALIASES.get(required_column, [required_column])
        actual_column = None
        for candidate_header in candidate_headers:
            normalized_required = _normalize_column_name(candidate_header)
            actual_column = normalized_to_actual.get(normalized_required)
            if actual_column is not None:
                break
        if actual_column is None:
            missing_columns.append(required_column)
            continue
        rename_map[actual_column] = required_column

    if missing_columns:
        raise ValueError(
            "Missing required source column(s): " + ", ".join(missing_columns)
        )

    return dataframe.rename(columns=rename_map)


def _load_encrypted_data(
    master_file_path: Path,
    sheet_name: str,
    master_password: str,
) -> pd.DataFrame:
    try:
        import msoffcrypto
    except ImportError as error:
        raise ValueError(
            "Password support requires 'msoffcrypto-tool'. Install it using: "
            "python -m pip install msoffcrypto-tool"
        ) from error

    decrypted_stream = BytesIO()
    try:
        with master_file_path.open("rb") as encrypted_file:
            office_file = msoffcrypto.OfficeFile(encrypted_file)
            office_file.load_key(password=master_password)
            office_file.decrypt(decrypted_stream)
    except Exception as error:
        raise ValueError(
            "Unable to decrypt master file. Check password and file format. "
            f"{error}"
        ) from error

    read_errors = []
    for engine in ("openpyxl", "xlrd"):
        decrypted_stream.seek(0)
        try:
            dataframe = pd.read_excel(
                decrypted_stream,
                sheet_name=sheet_name,
                dtype=object,
                engine=engine,
            )
            return dataframe
        except Exception as error:
            read_errors.append(f"{engine}: {error}")

    raise ValueError(
        "File was decrypted but sheet parsing failed. " + " | ".join(read_errors)
    )


def load_data(
    master_file_path: Path,
    sheet_name: str = SOURCE_SHEET,
    master_password: Optional[str] = None,
) -> pd.DataFrame:
    if not master_file_path.exists():
        raise FileNotFoundError(f"Master file not found: {master_file_path}")

    if master_password:
        dataframe = _load_encrypted_data(master_file_path, sheet_name, master_password)
    else:
        ole_signature = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        file_signature = master_file_path.read_bytes()[:8]

        if zipfile.is_zipfile(master_file_path):
            try:
                dataframe = pd.read_excel(
                    master_file_path,
                    sheet_name=sheet_name,
                    dtype=object,
                    engine="openpyxl",
                )
            except ValueError as error:
                raise ValueError(f"Unable to read sheet '{sheet_name}'. {error}")
        elif file_signature == ole_signature:
            try:
                dataframe = pd.read_excel(
                    master_file_path,
                    sheet_name=sheet_name,
                    dtype=object,
                    engine="xlrd",
                )
            except Exception as error:
                raise ValueError(
                    "Legacy/unsupported Excel container detected and could not be read. "
                    f"{error}. If this file is password-protected, enter workbook password. "
                    "Otherwise open in Excel and Save As true .xlsx."
                ) from error
        else:
            dataframe = None
            csv_errors = []
            for encoding in ("utf-8-sig", "latin-1"):
                try:
                    dataframe = pd.read_csv(
                        master_file_path,
                        dtype=object,
                        encoding=encoding,
                        sep=None,
                        engine="python",
                    )
                    break
                except Exception as csv_error:
                    csv_errors.append(f"{encoding}: {csv_error}")

            if dataframe is None:
                raise ValueError(
                    "Source file is not a valid .xlsx and CSV fallback failed. "
                    + " | ".join(csv_errors)
                )

    return _align_source_columns(dataframe)


def get_filter_requirements() -> Dict[str, object]:
    return {
        "column": "PLACEMENT",
        "condition": "in_list",
        "values": PLACEMENT_VALUES,
        "combine_logic": "AND",
        "case_sensitive": True,
        "keep_blank_rows": True,
    }


def filter_data(dataframe: pd.DataFrame, filter_requirements: Dict[str, object]) -> pd.DataFrame:
    column = str(filter_requirements.get("column", "")).strip()
    condition = str(filter_requirements.get("condition", "")).strip()
    values = filter_requirements.get("values", [])
    case_sensitive = bool(filter_requirements.get("case_sensitive", True))

    if column not in dataframe.columns:
        raise ValueError(f"Filter column not found: {column}")
    if condition != "in_list":
        raise ValueError(f"Invalid filter condition: {condition}")
    if not isinstance(values, list) or not values:
        raise ValueError("Invalid filter values: expected non-empty list")

    normalized_series = (
        dataframe[column]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    non_null_mask = dataframe[column].notna()

    if case_sensitive:
        value_set = {_normalize_placement_text(value) for value in values}
        mask = non_null_mask & normalized_series.isin(value_set)
    else:
        value_set = {_normalize_placement_text(value).casefold() for value in values}
        mask = non_null_mask & normalized_series.str.casefold().isin(value_set)

    filtered = dataframe[mask].copy()
    if filtered.empty:
        raise ValueError("Filtered result is empty. Check filter values and source data.")
    return filtered


def get_columns_to_extract() -> List[str]:
    return [
        "Product",
        "Bucket",
        "Account Number",
        "Mobile Number",
        "Email Address",
    ]


def _is_blank(value: object) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def extract_columns(
    dataframe: pd.DataFrame,
    columns_to_extract: List[str],
    date_of_blast: datetime,
) -> pd.DataFrame:
    missing_columns = [column for column in columns_to_extract if column not in dataframe.columns]
    if missing_columns:
        raise ValueError("Missing extraction column(s): " + ", ".join(missing_columns))

    output = pd.DataFrame(index=dataframe.index)
    for column in TEMPLATE_OUTPUT_COLUMNS:
        output[column] = ""

    for column in columns_to_extract:
        output[column] = dataframe[column]

    mobile_not_blank = output["Mobile Number"].notna() & output["Mobile Number"].astype(str).str.strip().ne("")
    email_not_blank = output["Email Address"].notna() & output["Email Address"].astype(str).str.strip().ne("")
    product_normalized = (
        output["Product"]
        .fillna("")
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    output.loc[mobile_not_blank, "SMS Dispo"] = "SMS - SENT"
    output.loc[mobile_not_blank, "IVR Dispo"] = "OTHER NON-VOICE - SENT"
    output.loc[email_not_blank, "Email Dispo"] = "EMAIL - SENT"
    output.loc[product_normalized.eq("Maya Easy Credit"), "Using Forbearance template ? Y/N"] = "Y"

    output["Agency"] = "SP MADRID"
    output["Date of Blast"] = date_of_blast.date()

    return output[TEMPLATE_OUTPUT_COLUMNS]


def write_to_template(output_dataframe: pd.DataFrame, sheet_name: str = OUTPUT_SHEET_NAME) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    headers = list(output_dataframe.columns)
    account_number_col_index = None
    date_of_blast_col_index = None
    if "Account Number" in headers:
        account_number_col_index = headers.index("Account Number") + 1
    if "Date of Blast" in headers:
        date_of_blast_col_index = headers.index("Date of Blast") + 1

    for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_index, value=header)

    for row_index, row_values in enumerate(output_dataframe.itertuples(index=False, name=None), start=2):
        for col_index, cell_value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if account_number_col_index is not None and col_index == account_number_col_index:
                cell.number_format = "@"
                cell.value = "" if pd.isna(cell_value) else str(cell_value)
            elif date_of_blast_col_index is not None and col_index == date_of_blast_col_index:
                cell.value = cell_value
                cell.number_format = "mm/dd/yyyy"
            else:
                cell.value = cell_value

    header_font = Font(name="Aptos", size=11, bold=True)
    body_font = Font(name="Aptos", size=11, bold=False)
    thin_black_border = Border(
        left=Side(style="thin", color="FF000000"),
        right=Side(style="thin", color="FF000000"),
        top=Side(style="thin", color="FF000000"),
        bottom=Side(style="thin", color="FF000000"),
    )

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for col_index in range(1, max_col + 1):
        header_value = worksheet.cell(row=1, column=col_index).value
        header_text = str(header_value) if header_value is not None else ""
        worksheet.cell(row=1, column=col_index).font = header_font
        for row_index in range(2, max_row + 1):
            worksheet.cell(row=row_index, column=col_index).font = body_font

        column_letter = get_column_letter(col_index)
        max_length = 0
        for row_index in range(1, max_row + 1):
            value = worksheet.cell(row=row_index, column=col_index).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        auto_width = max(max_length + OUTPUT_COLUMN_PADDING, OUTPUT_COLUMN_MIN_WIDTH)
        override_width = OUTPUT_COLUMN_WIDTH_OVERRIDES.get(header_text)
        final_width = override_width if override_width is not None else auto_width
        worksheet.column_dimensions[column_letter].width = min(final_width, OUTPUT_COLUMN_MAX_WIDTH)

    dispo_sheet = workbook.create_sheet(title=DIGITAL_DISPO_SHEET_NAME)
    header_cell = dispo_sheet.cell(row=1, column=1, value=DIGITAL_DISPO_COLUMN_NAME)
    header_cell.font = header_font
    header_cell.fill = PatternFill(
        fill_type="solid",
        start_color=DIGITAL_DISPO_HEADER_FILL,
        end_color=DIGITAL_DISPO_HEADER_FILL,
    )
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.border = thin_black_border

    for row_index, disposition in enumerate(DIGITAL_DISPO_VALUES, start=2):
        cell = dispo_sheet.cell(row=row_index, column=1, value=disposition)
        cell.font = body_font
        cell.border = thin_black_border

    max_dispo_length = max(len(DIGITAL_DISPO_COLUMN_NAME), *(len(value) for value in DIGITAL_DISPO_VALUES))
    dispo_sheet.column_dimensions["A"].width = min(max(max_dispo_length + 6, 34), OUTPUT_COLUMN_MAX_WIDTH)

    return workbook


def parse_date_from_filename(master_file_path: Path) -> datetime:
    filename_without_extension = master_file_path.stem
    date_match = re.search(r"(\d{8})", filename_without_extension)
    if not date_match:
        raise ValueError(
            "Unable to extract date from master filename. "
            "Expected MMDDYYYY in file name."
        )

    raw_date = date_match.group(1)
    try:
        return datetime.strptime(raw_date, "%m%d%Y")
    except ValueError as error:
        raise ValueError(
            f"Invalid date found in filename: {raw_date}. Expected MMDDYYYY format."
        ) from error


def process_master_file(master_path: Path, master_password: Optional[str] = None) -> pd.DataFrame:
    master_dataframe = load_data(
        master_path,
        sheet_name=SOURCE_SHEET,
        master_password=master_password,
    )
    filtered_dataframe = filter_data(master_dataframe, get_filter_requirements())
    transformed_dataframe = extract_columns(
        filtered_dataframe,
        get_columns_to_extract(),
        parse_date_from_filename(master_path),
    )
    return transformed_dataframe


st.set_page_config(page_title="Daily Digital", layout="wide")
st.title("Daily Digital")

with st.expander("Options", expanded=True):
    use_server_file = st.checkbox("Use server file by name", value=True)
    server_file_name = ""
    if use_server_file:
        st.caption(f"Server source: {DEFAULT_SERVER_MASTERFILE_DIR}")
        server_file_name = st.text_input(
            "Server file name",
            value=DEFAULT_SERVER_MASTERFILE_FILE,
            help="Paste filename, relative path (e.g. MARCH 2026\\file.xlsx), or full UNC file path.",
        )

wb_pass = st.text_input("Workbook Password", value=MF_WB_PASS, type="password")
uploaded_file = st.file_uploader(
    "Or Upload Masterfile",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=False,
)

if st.button("Process", type="primary", use_container_width=True):
    try:
        progress = st.progress(0, text="Starting...")
        uploaded_temp_file = False
        selected_file = None
        selected_name = None
        progress.progress(10, text="Locating source file...")

        if use_server_file:
            selected_path, server_err = resolve_server_masterfile(server_file_name)
            if server_err:
                st.error(server_err)
                st.stop()
            selected_file = selected_path
            selected_name = selected_path.name
            st.caption(f"Using server file: {selected_name}")
        elif uploaded_file is not None:
            selected_file = uploaded_file
            selected_name = uploaded_file.name
            st.caption(f"Using upload file: {selected_name}")
        else:
            st.error("Please upload a file or enable server file mode.")
            st.stop()

        if isinstance(selected_file, Path):
            target_path = selected_file
        else:
            temp_stream = BytesIO(selected_file.getvalue())
            target_path = Path.cwd() / selected_name
            target_path.write_bytes(temp_stream.getvalue())
            uploaded_temp_file = True
        progress.progress(35, text="Loading master data...")
        master_dataframe = load_data(
            target_path,
            sheet_name=SOURCE_SHEET,
            master_password=wb_pass.strip() or None,
        )

        date_of_blast = parse_date_from_filename(target_path)
        placement_counts_current = get_placement_counts(master_dataframe)
        placement_counts_comparison = None
        total_rows_comparison = None
        comparison_date = None

        kpi_cache = load_kpi_cache()
        placement_counts_comparison, total_rows_comparison, comparison_date = get_comparison_counts_from_cache(
            date_of_blast,
            kpi_cache,
        )

        if placement_counts_comparison is None:
            session_counts = st.session_state.get("digital_daily_last_counts")
            session_total_rows = st.session_state.get("digital_daily_last_total_rows")
            session_date = st.session_state.get("digital_daily_last_date")
            if isinstance(session_counts, dict) and isinstance(session_date, datetime):
                if session_date.date() < date_of_blast.date():
                    placement_counts_comparison = session_counts
                    if isinstance(session_total_rows, int):
                        total_rows_comparison = session_total_rows
                    comparison_date = session_date

        progress.progress(55, text="Filtering placements...")
        filtered_dataframe = filter_data(master_dataframe, get_filter_requirements())

        progress.progress(75, text="Transforming output columns...")
        transformed_dataframe = extract_columns(
            filtered_dataframe,
            get_columns_to_extract(),
            date_of_blast,
        )

        progress.progress(90, text="Building output workbook...")
        workbook = write_to_template(
            transformed_dataframe,
            sheet_name=OUTPUT_SHEET_NAME,
        )
        output_stream = BytesIO()
        workbook.save(output_stream)
        workbook.close()
        output_stream.seek(0)
        output_bytes = output_stream.getvalue()

        st.success(f"Processed {len(transformed_dataframe):,} rows.")

        st.subheader("Filter KPI")
        total_rows_current = len(transformed_dataframe)
        if total_rows_comparison is None and placement_counts_comparison is not None:
            total_rows_comparison = sum(placement_counts_comparison.get(placement, 0) for placement in PLACEMENT_VALUES)

        metric_columns = st.columns(5)
        total_rows_delta = None
        if total_rows_comparison is not None:
            total_rows_delta = f"{total_rows_current - total_rows_comparison:+,}"
        metric_columns[0].metric(
            label="Total Rows",
            value=f"{total_rows_current:,}",
            delta=total_rows_delta,
        )

        for index, placement in enumerate(PLACEMENT_VALUES):
            metric_label = PLACEMENT_METRIC_LABELS.get(placement, placement)
            metric_value = placement_counts_current.get(placement, 0)
            metric_delta = None
            if placement_counts_comparison is not None:
                base_value = placement_counts_comparison.get(placement, 0)
                metric_delta = f"{metric_value - base_value:+,}"
            metric_columns[index + 1].metric(
                label=metric_label,
                value=f"{metric_value:,}",
                delta=metric_delta,
            )

        if comparison_date is not None:
            if comparison_date.date() == (date_of_blast.date() - timedelta(days=1)):
                comparison_label = f"Compared to yesterday ({comparison_date.strftime('%m/%d/%Y')})"
            else:
                comparison_label = f"Compared to last used date ({comparison_date.strftime('%m/%d/%Y')})"
            st.caption(comparison_label)
        else:
            st.caption("No previous comparison file/date found yet.")

        st.dataframe(transformed_dataframe, use_container_width=True, height=400)

        output_filename = f"MADRID_DIGITAL RESULT_{date_of_blast.strftime('%Y-%m-%d')}.xlsx"
        progress.progress(95, text="Saving output to network folder...")
        try:
            saved_network_path = save_output_to_network(output_bytes, output_filename)
            st.caption(f"Saved to: {saved_network_path}")
        except FileExistsError as exists_error:
            st.warning(f"{exists_error}. File was not overwritten. Use the download button below.")
        except Exception as save_error:
            st.warning(f"Auto-save skipped: {save_error}")

        st.download_button(
            label="Download .xlsx",
            data=output_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        upsert_kpi_cache(date_of_blast, placement_counts_current, total_rows_current)

        st.session_state["digital_daily_last_counts"] = placement_counts_current
        st.session_state["digital_daily_last_total_rows"] = total_rows_current
        st.session_state["digital_daily_last_date"] = date_of_blast

        progress.progress(100, text="Done")
    except Exception as error:
        error_message = str(error).strip() or type(error).__name__
        st.error(f"Process failed: {error_message}")
        st.exception(error)
    finally:
        if (
            'target_path' in locals()
            and uploaded_temp_file
            and isinstance(target_path, Path)
            and target_path.exists()
        ):
            target_path.unlink(missing_ok=True)