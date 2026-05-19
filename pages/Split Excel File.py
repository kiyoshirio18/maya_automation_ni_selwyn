import streamlit as st
import openpyxl
from openpyxl import Workbook
from copy import copy
import os
import tempfile
import zipfile

def split_excel_with_formatting(input_file, original_filename, batch_size=500):
    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Get header
    header = [cell for cell in ws[1]]

    # Total rows (excluding header)
    total_rows = ws.max_row - 1
    num_batches = (total_rows // batch_size) + (1 if total_rows % batch_size else 0)

    # ✅ Use original filename (not temp name)
    base_name, ext = os.path.splitext(original_filename)

    output_files = []  # list of (path, filename)

    for batch in range(num_batches):
        # Create new workbook
        new_wb = Workbook()
        new_ws = new_wb.active

        # Copy header with formatting
        for col, cell in enumerate(header, start=1):
            new_cell = new_ws.cell(row=1, column=col, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        # Row range for this batch
        start_row = batch * batch_size + 2
        end_row = min(start_row + batch_size - 1, ws.max_row)

        # Copy rows with formatting
        for row_index, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row), start=2):
            for col, cell in enumerate(row, start=1):
                new_cell = new_ws.cell(row=row_index, column=col, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy column widths
        for col_letter, dim in ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width

        # ✅ Save temp file but with proper filename (original + _1, _2, etc.)
        temp_dir = tempfile.mkdtemp()
        filename = f"{base_name}_{batch+1}{ext}"  
        output_path = os.path.join(temp_dir, filename)
        new_wb.save(output_path)

        output_files.append((output_path, filename))

    return output_files


# ---------------- STREAMLIT UI ---------------- #

st.header("📑 Excel Splitter with Formatting")

uploaded_file = st.file_uploader("Upload an Excel file (.xlsx)", type=["xlsx"])
batch_size = st.number_input("Batch size (rows per file)", min_value=1, value=500, step=100)

if uploaded_file:
    st.success(f"Uploaded file: {uploaded_file.name}")

    if st.button("Split File"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name

        st.info("Processing... Please wait ⏳")
        output_files = split_excel_with_formatting(tmp_path, uploaded_file.name, batch_size=batch_size)

        # Debug log to confirm filenames
        for path, fname in output_files:
            st.write(f"Adding to ZIP → path: {path}, fname: {fname}")

        if len(output_files) == 1:
            path, fname = output_files[0]
            with open(path, "rb") as f:
                st.download_button(
                    label=f"📥 Download {fname}",
                    data=f,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            # Zip all files with correct names
            zip_path = tempfile.mktemp(suffix=".zip")
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for path, fname in output_files:
                    with open(path, "rb") as f:
                        zipf.writestr(fname, f.read())  # force correct filename inside ZIP

            with open(zip_path, "rb") as f:
                st.download_button(
                    label="📥 Download All Split Files (ZIP)",
                    data=f,
                    file_name=f"{os.path.splitext(uploaded_file.name)[0]}_split.zip",
                    mime="application/zip"
                )

        st.success("Done ✅")