import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

st.header("Contacted and Uncontacted Reports")
st.markdown("Generate cumulative reports by merging DRR data with account information.")
st.divider()

DEFAULT_MERGED_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\MERGED ACCOUNTS"
DEFAULT_REPORT_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\CONTACED AND UNCONTACTED"
TEMPLATE_PATH = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\CONTACED AND UNCONTACTED\template.xlsx"
REQUIRED_CYCLES = ["Maya Negosyo 121DPD", "Maya Negosyo 181DPD"]
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

FINAL_COLUMNS = [
    "DATE CALLED", "MONTH", "DCA NAME", "PRODUCT", "DPD", "BUCKET",
    "CPM ID", "ACCOUNT NUMBER", "MOBILE NUMBER CONTACTED", "DISPOSITION", "AMOUNT DUE"
]


def normalize_key(text: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(text).strip().upper()).strip("_")


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {normalize_key(col): col for col in df.columns if normalize_key(col)}
    for candidate in candidates:
        key = normalize_key(candidate)
        if key in lookup:
            return lookup[key]
    return None


def format_mobile_number(mobile) -> str:
    if pd.isna(mobile):
        return ""
    mobile_str = str(mobile).strip()
    
    if '.' in mobile_str:
        mobile_str = mobile_str.split('.')[0]
    
    mobile_str = ''.join(filter(str.isdigit, mobile_str))
    
    if len(mobile_str) >= 10:
        return f"63{mobile_str[-10:]}"
    return mobile_str


def get_bucket_from_cycle(cycle: str) -> str:
    cycle_upper = str(cycle).upper()
    if "121DPD" in cycle_upper:
        return "121 - 150 DPD"
    elif "181DPD" in cycle_upper:
        return "181 DPD & UP"
    return ""


def extract_date_from_filename(filename: str) -> str | None:
    patterns = [
        r"(\d{2})(\d{2})(\d{4})",  # MMDDYYYY
        r"(\d{4})-(\d{2})-(\d{2})",  # YYYY-MM-DD
        r"(\d{2})-(\d{2})-(\d{4})",  # MM-DD-YYYY
    ]
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            groups = match.groups()
            if len(groups[0]) == 4:  # YYYY-MM-DD
                return f"{groups[1]}{groups[2]}{groups[0]}"
            elif len(groups[2]) == 4:  # MM-DD-YYYY or MMDDYYYY
                return f"{groups[0]}{groups[1]}{groups[2]}"
    return None


def get_latest_file(directory: str, pattern: str = "*.xlsx") -> Path | None:
    dir_path = Path(directory)
    if not dir_path.exists() or not dir_path.is_dir():
        return None
    
    excel_files = [f for f in dir_path.glob(pattern) if not f.name.startswith("~$") and f.suffix.lower() in ALLOWED_EXTENSIONS]
    if not excel_files:
        return None
    
    return max(excel_files, key=lambda f: f.stat().st_mtime)


def resolve_merged_file(merged_input: str) -> tuple[Path | None, str | None]:
    from datetime import date
    
    root_dir = Path(DEFAULT_MERGED_DIR)
    if not root_dir.exists() or not root_dir.is_dir():
        return None, f"Merged accounts folder is not reachable: {root_dir}"
    
    requested = (merged_input or "").strip()
    if not requested:
        return None, "Please provide a file name or relative path."
    
    normalized = requested.replace("/", "\\")
    requested_path = Path(normalized)
    
    if requested_path.exists() and requested_path.is_file():
        return requested_path, None
    
    direct_candidate = root_dir / normalized
    if direct_candidate.exists() and direct_candidate.is_file():
        return direct_candidate, None
    
    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid file name or relative path."
    
    current_year_dir = root_dir / str(date.today().year)
    search_roots = [current_year_dir, root_dir] if current_year_dir.exists() else [root_dir]
    
    has_extension = Path(requested_name).suffix.lower() in ALLOWED_EXTENSIONS
    matches: list[Path] = []
    
    for base in search_roots:
        if not base.exists() or not base.is_dir():
            continue
        
        for file_path in base.rglob("*"):
            if not file_path.is_file() or file_path.suffix.lower() not in ALLOWED_EXTENSIONS:
                continue
            
            if has_extension:
                if file_path.name.lower() == requested_name.lower():
                    matches.append(file_path)
            else:
                if file_path.stem.lower() == requested_name.lower() or file_path.name.lower() == requested_name.lower():
                    matches.append(file_path)
    
    unique_matches = sorted(set(matches))
    if len(unique_matches) == 1:
        return unique_matches[0], None
    
    if len(unique_matches) > 1:
        preview = "\n".join(str(p) for p in unique_matches[:8])
        return None, (
            f"Multiple files matched '{requested_name}'. Please provide a more specific relative path.\n"
            f"Sample matches:\n{preview}"
        )
    
    return None, f"File not found from input: {requested}"


def classify_status(status: str) -> str:
    status_upper = str(status).upper()
    
    if "PAYMENT" in status_upper:
        return "EXCLUDE"
    
    if "POSITIVE" in status_upper and "CONTACT" in status_upper:
        return "CONTACTED"
    
    if "PTP" in status_upper or "CONTACTED" in status_upper:
        return "CONTACTED"
    
    return "UNCONTACTED"


def load_and_filter_drr(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    
    cycle_col = find_column(df, ["CYCLE"])
    status_col = find_column(df, ["STATUS", "CALL STATUS", "DISPOSITION"])
    acct_col = find_column(df, ["ACCOUNT NUMBER", "ACCOUNT NO", "ACCOUNT_NUMBER", "ACCT_NO"])
    
    if cycle_col is None:
        raise ValueError("Missing required column: CYCLE")
    if status_col is None:
        raise ValueError("Missing required column: STATUS")
    if acct_col is None:
        raise ValueError("Missing required column: ACCOUNT NUMBER")
    
    df = df[df[cycle_col].isin(REQUIRED_CYCLES)].copy()
    df["_CLASSIFICATION"] = df[status_col].apply(classify_status)
    df = df[df["_CLASSIFICATION"] != "EXCLUDE"].copy()
    
    return df


def enrich_with_merged_data(drr_df: pd.DataFrame, merged_path: Path) -> pd.DataFrame:
    suffix = merged_path.suffix.lower()
    if suffix == ".csv":
        merged_df = pd.read_csv(merged_path)
    else:
        merged_df = pd.read_excel(merged_path)
    
    drr_acct_col = find_column(drr_df, ["ACCOUNT NUMBER", "ACCOUNT NO", "ACCOUNT_NUMBER", "ACCT_NO"])
    merged_acct_col = find_column(merged_df, ["ACCOUNT NUMBER", "ACCOUNT NO", "ACCOUNT_NUMBER", "ACCT_NO"])
    cpm_col = find_column(merged_df, ["CPM ID", "CPM_ID", "CPMID"])
    mobile_col = find_column(merged_df, ["MOBILE PROPER", "MOBILE NUMBER", "MOBILE_NUMBER", "MOBILE"])
    amount_col = find_column(merged_df, ["AMOUNT DUE", "AMOUNT_DUE", "AMOUNTDUE", "BALANCE"])
    
    if not all([drr_acct_col, merged_acct_col]):
        raise ValueError("Missing ACCOUNT NUMBER column in DRR or merged file")
    
    lookup_cols = [merged_acct_col]
    if cpm_col:
        lookup_cols.append(cpm_col)
    if mobile_col:
        lookup_cols.append(mobile_col)
    if amount_col:
        lookup_cols.append(amount_col)
    
    merged_subset = merged_df[lookup_cols].drop_duplicates(subset=[merged_acct_col])
    
    enriched = drr_df.merge(
        merged_subset,
        left_on=drr_acct_col,
        right_on=merged_acct_col,
        how="left"
    )
    
    if mobile_col and mobile_col in enriched.columns:
        enriched["MOBILE NUMBER CONTACTED"] = enriched[mobile_col].apply(format_mobile_number)
    
    return enriched


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    col_mapping = {}
    
    special_mappings = {
        "DATE CALLED": ["DATE CALLED", "DATE"],
        "DCA NAME": ["DCA NAME", "COLLECTOR"],
        "PRODUCT": ["PRODUCT", "PRODUCT DESCRIPTION", "PRODUCT TYPE"],
        "ACCOUNT NUMBER": ["ACCOUNT NUMBER", "ACCOUNT NO", "ACCOUNT_NUMBER"],
        "DISPOSITION": ["DISPOSITION", "PROPOSED DISPOSITION"],
        "AMOUNT DUE": ["AMOUNT DUE", "BALANCE", "AMOUNT_DUE"],
        "CYCLE": ["CYCLE"]
    }
    
    for target_col in FINAL_COLUMNS:
        candidates = special_mappings.get(target_col, [target_col])
        found_col = find_column(df, candidates)
        if found_col and found_col != target_col:
            col_mapping[found_col] = target_col
    
    cycle_col = find_column(df, ["CYCLE"])
    if cycle_col:
        col_mapping[cycle_col] = "_CYCLE_TEMP"
    
    df = df.rename(columns=col_mapping)
    
    for col in FINAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    
    if "DATE CALLED" in df.columns:
        df["DATE CALLED"] = pd.to_datetime(df["DATE CALLED"], errors="coerce")
        df["MONTH"] = df["DATE CALLED"].dt.strftime("%b-%y")
        df["DATE CALLED"] = df["DATE CALLED"].dt.date
    
    df["DCA NAME"] = "MADRID"
    df["PRODUCT"] = "MAYA NEGOSYO ADVANCED"
    
    if "_CYCLE_TEMP" in df.columns:
        df["BUCKET"] = df["_CYCLE_TEMP"].apply(get_bucket_from_cycle)
        df = df.drop(columns=["_CYCLE_TEMP"])
    
    if "ACCOUNT NUMBER" in df.columns:
        df["ACCOUNT NUMBER"] = pd.to_numeric(df["ACCOUNT NUMBER"], errors="coerce").fillna(0).astype(int)
    
    if "MOBILE NUMBER CONTACTED" in df.columns:
        df["MOBILE NUMBER CONTACTED"] = df["MOBILE NUMBER CONTACTED"].astype(str)
    
    if "AMOUNT DUE" in df.columns:
        df["AMOUNT DUE"] = pd.to_numeric(df["AMOUNT DUE"], errors="coerce").fillna(0).round(2)
    
    return df[FINAL_COLUMNS]


def append_and_deduplicate(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    if not old_df.empty:
        if "DATE CALLED" in old_df.columns:
            old_df["DATE CALLED"] = pd.to_datetime(old_df["DATE CALLED"], errors="coerce").dt.date
        
        if "MONTH" in old_df.columns:
            temp_date = pd.to_datetime(old_df["DATE CALLED"], errors="coerce")
            old_df["MONTH"] = temp_date.dt.strftime("%b-%y")
        
        if "MOBILE NUMBER CONTACTED" in old_df.columns:
            old_df["MOBILE NUMBER CONTACTED"] = old_df["MOBILE NUMBER CONTACTED"].astype(str).str.replace(".0", "", regex=False)
        
        if "ACCOUNT NUMBER" in old_df.columns:
            old_df["ACCOUNT NUMBER"] = pd.to_numeric(old_df["ACCOUNT NUMBER"], errors="coerce").fillna(0).astype(int)
        
        if "AMOUNT DUE" in old_df.columns:
            old_df["AMOUNT DUE"] = pd.to_numeric(old_df["AMOUNT DUE"], errors="coerce").fillna(0).round(2)
    
    combined = pd.concat([old_df, new_df], ignore_index=True)
    
    date_col = "DATE CALLED"
    acct_col = "ACCOUNT NUMBER"
    
    if date_col in combined.columns:
        combined[date_col] = pd.to_datetime(combined[date_col], errors="coerce").dt.date
    
    combined = combined.drop_duplicates(subset=[acct_col, date_col], keep="last")
    
    return combined


def generate_report(drr_file, merged_path: Path, report_dir: str, template_path: str):
    with st.status("Loading DRR file...", expanded=True) as status:
        drr_df = load_and_filter_drr(drr_file)
        
        if drr_df.empty:
            st.warning("No records found matching the cycle filter.")
            status.update(label="No matching records", state="error")
            return
        
        st.success(f"Loaded {len(drr_df)} records")
        status.update(label="DRR file loaded", state="complete")
    
    with st.status(f"Enriching with merged data...", expanded=True) as status:
        enriched_df = enrich_with_merged_data(drr_df, merged_path)
        contacted_df = enriched_df[enriched_df["_CLASSIFICATION"] == "CONTACTED"].copy()
        uncontacted_df = enriched_df[enriched_df["_CLASSIFICATION"] == "UNCONTACTED"].copy()
        
        contacted_df = normalize_columns(contacted_df)
        uncontacted_df = normalize_columns(uncontacted_df)
        
        st.success(f"CONTACTED: {len(contacted_df)} | UNCONTACTED: {len(uncontacted_df)}")
        status.update(label="Data enriched", state="complete")
    
    with st.status("Loading existing report...", expanded=True) as status:
        latest_report = get_latest_file(report_dir)
        
        if latest_report:
            st.info(f"Found: {latest_report.name}")
            old_contacted = pd.read_excel(latest_report, sheet_name="CONTACTED")
            old_uncontacted = pd.read_excel(latest_report, sheet_name="UNCONTACTED")
            status.update(label="Existing report loaded", state="complete")
        else:
            st.info("No existing report. Using template.")
            template = Path(template_path)
            if not template.exists():
                status.update(label="Template not found", state="error")
                raise FileNotFoundError(f"Template file not found: {template_path}")
            old_contacted = pd.read_excel(template, sheet_name="CONTACTED")
            old_uncontacted = pd.read_excel(template, sheet_name="UNCONTACTED")
            status.update(label="Template loaded", state="complete")
    
    with st.status("Appending and deduplicating...", expanded=True) as status:
        final_contacted = append_and_deduplicate(old_contacted, contacted_df)
        final_uncontacted = append_and_deduplicate(old_uncontacted, uncontacted_df)
        st.success(f"Final - CONTACTED: {len(final_contacted)} | UNCONTACTED: {len(final_uncontacted)}")
        status.update(label="Data deduplicated", state="complete")
    
    date_str = extract_date_from_filename(drr_file.name)
    if not date_str:
        date_str = datetime.now().strftime("%m%d%Y")
    
    output_filename = f"CONANDUNCON_{date_str}-update.xlsx"
    output_path = Path(report_dir) / output_filename
    
    with st.status(f"Saving report: {output_filename}", expanded=True) as status:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            final_contacted.to_excel(writer, sheet_name="CONTACTED", index=False)
            final_uncontacted.to_excel(writer, sheet_name="UNCONTACTED", index=False)
            
            workbook = writer.book
            for sheet_name in ["CONTACTED", "UNCONTACTED"]:
                worksheet = writer.sheets[sheet_name]
                
                for col_idx, col_name in enumerate(FINAL_COLUMNS, start=1):
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    
                    if col_name == "ACCOUNT NUMBER":
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.number_format = '0'
                    
                    elif col_name == "MOBILE NUMBER CONTACTED":
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.number_format = '@'
                    
                    elif col_name == "AMOUNT DUE":
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.number_format = '0.00'
                    
                    max_length = len(str(col_name))
                    for row in range(2, worksheet.max_row + 1):
                        cell_value = worksheet.cell(row=row, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    
                    adjusted_width = min(max_length + 5, 60)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
            
            if latest_report:
                try:
                    from openpyxl import load_workbook
                    source_wb = load_workbook(latest_report)
                    if "instruction" in source_wb.sheetnames:
                        source_sheet = source_wb["instruction"]
                        target_sheet = workbook.create_sheet("instruction")
                        
                        for row in source_sheet.iter_rows():
                            for cell in row:
                                target_cell = target_sheet[cell.coordinate]
                                target_cell.value = cell.value
                                if cell.has_style:
                                    target_cell.font = cell.font.copy()
                                    target_cell.border = cell.border.copy()
                                    target_cell.fill = cell.fill.copy()
                                    target_cell.number_format = cell.number_format
                                    target_cell.protection = cell.protection.copy()
                                    target_cell.alignment = cell.alignment.copy()
                        
                        for col_letter, col_dim in source_sheet.column_dimensions.items():
                            target_sheet.column_dimensions[col_letter].width = col_dim.width
                        
                        for row_num, row_dim in source_sheet.row_dimensions.items():
                            target_sheet.row_dimensions[row_num].height = row_dim.height
                except:
                    pass
            else:
                template = Path(template_path)
                if template.exists():
                    try:
                        from openpyxl import load_workbook
                        template_wb = load_workbook(template)
                        if "instruction" in template_wb.sheetnames:
                            source_sheet = template_wb["instruction"]
                            target_sheet = workbook.create_sheet("instruction")
                            
                            for row in source_sheet.iter_rows():
                                for cell in row:
                                    target_cell = target_sheet[cell.coordinate]
                                    target_cell.value = cell.value
                                    if cell.has_style:
                                        target_cell.font = cell.font.copy()
                                        target_cell.border = cell.border.copy()
                                        target_cell.fill = cell.fill.copy()
                                        target_cell.number_format = cell.number_format
                                        target_cell.protection = cell.protection.copy()
                                        target_cell.alignment = cell.alignment.copy()
                            
                            for col_letter, col_dim in source_sheet.column_dimensions.items():
                                target_sheet.column_dimensions[col_letter].width = col_dim.width
                            
                            for row_num, row_dim in source_sheet.row_dimensions.items():
                                target_sheet.row_dimensions[row_num].height = row_dim.height
                    except:
                        pass
        
        st.success(f"Saved to: {output_path}")
        status.update(label="Report saved successfully", state="complete")
    
    st.balloons()
    st.success(f"**Report generated successfully!**")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("CONTACTED", f"{len(final_contacted):,}")
    with col2:
        st.metric("UNCONTACTED", f"{len(final_uncontacted):,}")
    with col3:
        st.metric("TOTAL", f"{len(final_contacted) + len(final_uncontacted):,}")
    
    st.info(f"**File:** `{output_filename}`")


with st.form("contacted_uncontacted_form"):
    st.subheader("Input Files")
    
    drr_uploaded = st.file_uploader(
        "1. Upload Processed DRR File",
        type=["xlsx"],
        help="Upload the processed DRR file containing STATUS and CYCLE columns",
        key="drr_file"
    )
    
    st.divider()
    st.subheader("Merged Account File")
    
    from datetime import date
    default_merged_name = f"maya_merged_accounts_{date.today().strftime('%m%d%y')}"
    
    col1, col2 = st.columns([1, 3])
    with col1:
        use_server_merged = st.checkbox("Use server file", value=True, key="use_server_merged")
    with col2:
        if use_server_merged:
            st.caption(f"Base: `{DEFAULT_MERGED_DIR}`")
    
    if use_server_merged:
        merged_filename = st.text_input(
            "File name or relative path",
            value=default_merged_name,
            placeholder=default_merged_name,
            help="Example: maya_merged_accounts_041626 or 2026\\APRIL\\maya_merged_accounts_041626.xlsx",
            key="merged_filename"
        )
    else:
        merged_uploaded = st.file_uploader(
            "Upload merged account file",
            type=["xlsx", "xls", "csv"],
            help="Upload the merged account file manually",
            key="merged_upload"
        )
    
    st.divider()
    st.subheader("Output Settings")
    
    col3, col4 = st.columns([1, 3])
    with col3:
        use_default_report_dir = st.checkbox("Use default directory", value=True, key="use_default_dir")
    with col4:
        if use_default_report_dir:
            st.caption(f"Output: `{DEFAULT_REPORT_DIR}`")
    
    if not use_default_report_dir:
        report_directory = st.text_input(
            "Custom report directory path",
            value=DEFAULT_REPORT_DIR,
            help="Directory where reports are stored and will be saved",
            key="report_dir"
        )
    else:
        report_directory = DEFAULT_REPORT_DIR
    
    st.divider()
    submitted = st.form_submit_button("Generate Report", use_container_width=True, type="primary")

if submitted:
    if not drr_uploaded:
        st.error("Please upload a DRR file to continue.")
    else:
        with st.spinner("Processing report..."):
            try:
                selected_merged_file = None
                
                if use_server_merged:
                    with st.status("Resolving merged account file...", expanded=False) as status:
                        selected_merged_file, merged_error = resolve_merged_file(merged_filename)
                        if merged_error:
                            st.warning(merged_error)
                            status.update(label="Server file not found", state="error")
                        else:
                            st.success(f"Found: {selected_merged_file.name}")
                            status.update(label="Merged file loaded", state="complete")
                else:
                    if 'merged_uploaded' in locals() and merged_uploaded is not None:
                        with st.status("Loading uploaded merged file...", expanded=False) as status:
                            import tempfile
                            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(merged_uploaded.name).suffix) as tmp:
                                tmp.write(merged_uploaded.read())
                                selected_merged_file = Path(tmp.name)
                            st.success(f"Loaded: {merged_uploaded.name}")
                            status.update(label="Merged file loaded", state="complete")
                
                if selected_merged_file is None:
                    st.error("No merged account file available. Please provide a valid file.")
                else:
                    generate_report(drr_uploaded, selected_merged_file, report_directory, TEMPLATE_PATH)
                
            except Exception as exc:
                st.error(f"Processing failed: {exc}")
                st.exception(exc)
