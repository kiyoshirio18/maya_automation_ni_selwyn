import streamlit as st
import polars as pl
import msoffcrypto
from io import BytesIO
from openpyxl import Workbook

st.header("Split Masterfile Per Placement")
st.write(f''':green[:green-background[PLACEMENT]]''')
output = None
date = None

with st.form(key="placement"):
    file = st.file_uploader("Upload Password Protected XLSX", type=["xlsx"])

    if file is not None:
        filename = file.name
        date = (filename.rpartition('.')[0]).split()[-1]

    if st.form_submit_button("SUBMIT", use_container_width=True):
        # if file is not None:
        #     excel_decrypted = "./temp/decrypted.xlsx"
        #     with BytesIO(file.read()) as f:
        #         excel_file = msoffcrypto.OfficeFile(f)
        #         excel_file.load_key(password = passkey)
        #         with open(excel_decrypted, 'wb') as df:
        #             excel_file.decrypt(df)

        masterfile_df = pl.read_excel(file)
        placements = masterfile_df["PLACEMENT"].unique().sort().to_list()
        wb = Workbook()

        for value in placements:
            # Filter the DataFrame based on the unique value
            filtered_df = masterfile_df.filter(pl.col("PLACEMENT") == value)
            
            # Create a new sheet in the workbook
            short_name = value.lstrip("Maya ")

            ws = wb.create_sheet(title=short_name)

            ws.append(filtered_df.columns)
            
            # Write the DataFrame to the sheet
            for row in filtered_df.iter_rows(named=True):
                ws.append(list(row.values()))

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("Sheet"):
                wb.remove(wb[sheet_name])
       
        output = BytesIO()
        wb.save(output)
        output.seek(0)

if output is not None:
    st.download_button(
        label = "Download .xlsx",
        data = output,
        file_name = f"maya_per_placement_{date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )